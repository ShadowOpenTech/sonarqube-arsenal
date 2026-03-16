#!/usr/bin/env python3
"""
SonarQube Trend Report
======================
On-demand report script.  Reads N most recent snapshot JSON files from
--snapshots-dir, loads a department mapping CSV, and produces a multi-sheet
Excel trend report with line charts.

No SonarQube API calls are made — this script only reads local files.

Sheets produced:
  Summary sheet (first)  — inserted at position 0
  Fleet Trend            — active/acknowledged vuln totals over time, with chart
  Dept - <Name>          — one sheet per dept (active vulns per CODE, with chart)
  Code - <Name>          — one sheet per CODE  (active vulns per project, with chart)
  Events                 — branch changes, added/removed projects, unmapped codes
  Coverage Gaps          — inactive projects and unmapped codes

Usage:
    python sonar_trend_report.py \\
        --mapping dept_mapping.csv \\
        --snapshots-dir snapshots \\
        --output trend_report.xlsx

    # Limit to last 12 snapshots (quarterly view)
    python sonar_trend_report.py \\
        --mapping dept_mapping.csv \\
        --snapshots-count 12 \\
        --output trend_report.xlsx

    # Only internet-facing, critical projects
    python sonar_trend_report.py \\
        --mapping dept_mapping.csv \\
        --internet-facing-only \\
        --critical-only

Requirements:
    pip install openpyxl tqdm python-dotenv
"""

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ──────────────────────────────────────────────────────────────────

DEFAULT_SNAPSHOTS_COUNT = 52
MAX_CHART_SERIES        = 20   # max lines per chart to avoid illegibility
CHART_WIDTH             = 30
CHART_HEIGHT            = 15


# ── Styles ─────────────────────────────────────────────────────────────────────

_HDR_FILL      = PatternFill("solid", fgColor="1F3864")   # dark navy
_TOTAL_FILL    = PatternFill("solid", fgColor="375623")   # dark green
_ALT_FILL      = PatternFill("solid", fgColor="DEEAF1")   # light blue alternating
_WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
_YELLOW_FILL   = PatternFill("solid", fgColor="FFFF99")   # internet-facing
_RED_FILL      = PatternFill("solid", fgColor="FFCCCC")   # critical
_BLUE_FILL     = PatternFill("solid", fgColor="BDD7EE")   # branch changed
_GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")   # added
_ORANGE_FILL   = PatternFill("solid", fgColor="FCE4D6")   # removed / any-stale
_LORANGE_FILL  = PatternFill("solid", fgColor="FFDBB3")   # stale / partial

_HDR_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT     = Font(name="Calibri", size=10)
_BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
_TOTAL_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

_THIN          = Side(border_style="thin", color="B8CCE4")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center")
_RIGHT         = Alignment(horizontal="right",  vertical="center")


# ── Low-level cell helpers ─────────────────────────────────────────────────────

def _hdr(ws, row: int, col: int, value: str, fill=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill or _HDR_FILL
    c.font      = _HDR_FONT
    c.alignment = _CENTER
    c.border    = _BORDER


def _cell(ws, row: int, col: int, value,
          fill=None, align=None, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill or _WHITE_FILL
    c.font      = _BOLD_FONT if bold else _BODY_FONT
    c.alignment = align or _LEFT
    c.border    = _BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


# ── Data loading ───────────────────────────────────────────────────────────────

def load_snapshots(snapshots_dir: str, count: int) -> list[dict]:
    """Load the `count` most recent snapshot JSON files, sorted chronologically."""
    p = Path(snapshots_dir)
    files = sorted(p.glob("snapshot_*.json"))
    files = files[-count:]
    snapshots: list[dict] = []
    for f in tqdm(files, desc="Loading snapshots", unit="file"):
        try:
            with open(f, encoding="utf-8") as fh:
                snapshots.append(json.load(fh))
        except Exception as exc:
            tqdm.write(f"  [WARN] Could not read snapshot {f.name}: {exc}",
                       file=sys.stderr)
    return snapshots


def load_dept_mapping(path: str) -> dict[str, dict]:
    """Load dept mapping CSV.  Returns {code: {dept, internet_facing, critical}}."""
    mapping: dict[str, dict] = {}
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row.get("CODE", "").strip()
            if not code:
                continue
            mapping[code] = {
                "dept":            row.get("DEPT", "").strip(),
                "internet_facing": row.get("internetFacing", "false").strip().lower() == "true",
                "critical":        row.get("Critical", "false").strip().lower() == "true",
            }
    return mapping


def load_project_keys_file(path: str) -> list[str]:
    with open(path, encoding="utf-8") as f:
        return [
            line.strip()
            for line in f
            if line.strip() and not line.strip().startswith("#")
        ]


def _extract_code(name: str) -> str:
    return name.split("_")[0] if "_" in name else name


def _snap_date(snapshot: dict) -> str:
    """Return YYYY-MM-DD from snapshot generated_at."""
    iso = snapshot.get("generated_at", "")
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return snapshot.get("snapshot_id", "unknown")[:10]


def _active_vulns(project: dict) -> int:
    v = project.get("vulnerabilities")
    if not v:
        return 0
    return v.get("active", 0)


def _acknowledged_vulns(project: dict) -> int:
    v = project.get("vulnerabilities")
    if not v:
        return 0
    return v.get("acknowledged", 0)


def _to_review_hotspots(project: dict) -> int:
    h = project.get("hotspots")
    if not h:
        return 0
    return h.get("to_review", 0)


# ── Time-series builders ───────────────────────────────────────────────────────

def build_fleet_series(snapshots: list[dict]) -> dict:
    """Return {dates: [...], active: [...], acknowledged: [...], to_review: [...]}."""
    dates: list[str] = []
    active_series:       list[int] = []
    acknowledged_series: list[int] = []
    to_review_series:    list[int] = []

    for snap in snapshots:
        dates.append(_snap_date(snap))
        total_active = total_ack = total_tr = 0
        for proj in snap.get("projects", []):
            if proj.get("error"):
                continue
            total_active += _active_vulns(proj)
            total_ack    += _acknowledged_vulns(proj)
            total_tr     += _to_review_hotspots(proj)
        active_series.append(total_active)
        acknowledged_series.append(total_ack)
        to_review_series.append(total_tr)

    return {
        "dates":        dates,
        "active":       active_series,
        "acknowledged": acknowledged_series,
        "to_review":    to_review_series,
    }


def build_dept_series(
    snapshots: list[dict],
    dept_mapping: dict[str, dict],
) -> dict[str, dict]:
    """
    Returns {dept_name: {dates, codes: {code: [values]}, totals: [values]}}.
    """
    dates = [_snap_date(s) for s in snapshots]
    # dept → code → list of values per snapshot
    dept_code_values: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    dept_totals: dict[str, list[int]] = defaultdict(list)

    for snap in snapshots:
        # accumulate per dept/code for this snapshot
        snap_dept_code: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for proj in snap.get("projects", []):
            if proj.get("error"):
                continue
            code = proj.get("code") or _extract_code(proj.get("name", proj.get("key", "")))
            entry = dept_mapping.get(code, {})
            dept = entry.get("dept", "Unmapped")
            snap_dept_code[dept][code] += _active_vulns(proj)

        # append values; ensure codes that don't appear get 0
        all_depts = set(dept_code_values.keys()) | set(snap_dept_code.keys())
        for dept in all_depts:
            all_codes = set(dept_code_values[dept].keys()) | set(snap_dept_code[dept].keys())
            dept_total = 0
            for code in all_codes:
                val = snap_dept_code[dept].get(code, 0)
                dept_code_values[dept][code].append(val)
                dept_total += val
            dept_totals[dept].append(dept_total)

    result: dict[str, dict] = {}
    for dept in dept_code_values:
        result[dept] = {
            "dates":  dates,
            "codes":  dict(dept_code_values[dept]),
            "totals": dept_totals[dept],
        }
    return result


def build_code_series(
    snapshots: list[dict],
) -> dict[str, dict]:
    """
    Returns {code: {dates, projects: {project_key: [values]}}}.
    """
    dates = [_snap_date(s) for s in snapshots]
    code_proj_values: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))

    for snap in snapshots:
        snap_code_proj: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for proj in snap.get("projects", []):
            if proj.get("error"):
                continue
            key  = proj.get("key", "")
            code = proj.get("code") or _extract_code(proj.get("name", key))
            snap_code_proj[code][key] += _active_vulns(proj)

        all_codes = set(code_proj_values.keys()) | set(snap_code_proj.keys())
        for code in all_codes:
            all_projs = set(code_proj_values[code].keys()) | set(snap_code_proj[code].keys())
            for proj_key in all_projs:
                val = snap_code_proj[code].get(proj_key, 0)
                code_proj_values[code][proj_key].append(val)

    result: dict[str, dict] = {}
    for code in code_proj_values:
        result[code] = {
            "dates":    dates,
            "projects": dict(code_proj_values[code]),
        }
    return result


# ── Event detection ────────────────────────────────────────────────────────────

def detect_events(
    snapshots: list[dict],
    dept_mapping: dict[str, dict],
) -> list[dict]:
    """Scan snapshots for notable events.  Returns list of event dicts."""
    events: list[dict] = []

    prev_keys: Optional[set] = None

    for snap in snapshots:
        date = _snap_date(snap)
        curr_keys: set[str] = set()

        if snap.get("status") == "partial":
            events.append({
                "date":        date,
                "event_type":  "Snapshot Partial",
                "project_key": "",
                "project_name": "",
                "code":        "",
                "dept":        "",
                "detail":      f"Snapshot {snap.get('snapshot_id', '')} has status=partial "
                               f"({snap.get('projects_failed', 0)} project(s) failed)",
            })

        for proj in snap.get("projects", []):
            key  = proj.get("key", "")
            name = proj.get("name", key)
            code = proj.get("code") or _extract_code(name)
            dept = dept_mapping.get(code, {}).get("dept", "")

            if proj.get("error") and not proj["error"].startswith("skipped"):
                # Don't generate events for API errors — they're logged during snapshot
                pass

            if proj.get("branch_changed"):
                events.append({
                    "date":         date,
                    "event_type":   "Branch Changed",
                    "project_key":  key,
                    "project_name": name,
                    "code":         code,
                    "dept":         dept,
                    "detail":       f"Branch changed to {proj.get('branch', '')}",
                })

            if not proj.get("error"):
                curr_keys.add(key)

            # Unmapped code
            if code not in dept_mapping:
                events.append({
                    "date":         date,
                    "event_type":   "Unmapped Code",
                    "project_key":  key,
                    "project_name": name,
                    "code":         code,
                    "dept":         "",
                    "detail":       f"Code '{code}' not found in dept mapping",
                })

            # Stale branch warning
            age = proj.get("branch_age_days")
            if age is not None and age > 60 and not proj.get("error"):
                events.append({
                    "date":         date,
                    "event_type":   "Branch Stale",
                    "project_key":  key,
                    "project_name": name,
                    "code":         code,
                    "dept":         dept,
                    "detail":       f"Branch last analysed {age} days ago",
                })

        if prev_keys is not None:
            for k in curr_keys - prev_keys:
                proj_data = next(
                    (p for p in snap.get("projects", []) if p.get("key") == k), {}
                )
                name = proj_data.get("name", k)
                code = proj_data.get("code") or _extract_code(name)
                events.append({
                    "date":         date,
                    "event_type":   "Project Added",
                    "project_key":  k,
                    "project_name": name,
                    "code":         code,
                    "dept":         dept_mapping.get(code, {}).get("dept", ""),
                    "detail":       "Project appeared in this snapshot",
                })
            for k in prev_keys - curr_keys:
                events.append({
                    "date":         date,
                    "event_type":   "Project Removed",
                    "project_key":  k,
                    "project_name": k,
                    "code":         _extract_code(k),
                    "dept":         "",
                    "detail":       "Project disappeared from this snapshot",
                })

        prev_keys = curr_keys

    return events


# ── Apply filters ──────────────────────────────────────────────────────────────

def _filter_projects_in_snapshots(
    snapshots: list[dict],
    filter_keys: Optional[set[str]],
    dept_mapping: dict[str, dict],
    internet_facing_only: bool,
    critical_only: bool,
) -> list[dict]:
    """Return snapshots with project lists filtered per CLI flags."""
    filtered: list[dict] = []
    for snap in snapshots:
        new_projects: list[dict] = []
        for proj in snap.get("projects", []):
            key  = proj.get("key", "")
            code = proj.get("code") or _extract_code(proj.get("name", key))
            entry = dept_mapping.get(code, {})

            if filter_keys and key not in filter_keys:
                continue
            if internet_facing_only and not entry.get("internet_facing", False):
                continue
            if critical_only and not entry.get("critical", False):
                continue
            new_projects.append(proj)

        new_snap = dict(snap)
        new_snap["projects"] = new_projects
        filtered.append(new_snap)
    return filtered


# ── Chart helper ───────────────────────────────────────────────────────────────

def _add_line_chart(ws, title: str, anchor_cell: str,
                    data_ref: Reference, cats_ref: Reference,
                    series_titles: Optional[list[str]] = None) -> None:
    chart = LineChart()
    chart.title  = title
    chart.style  = 10
    chart.width  = CHART_WIDTH
    chart.height = CHART_HEIGHT
    chart.y_axis.title = "Active Vulnerabilities"
    chart.x_axis.title = "Date"

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, anchor_cell)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_overview_sheet(ws, snapshots: list[dict], dept_mapping: dict) -> None:
    ws.title = "Overview"
    ws.freeze_panes = "A2"

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 28

    latest = snapshots[-1] if snapshots else {}
    rows = [
        ("Report generated at",     datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Snapshots loaded",         len(snapshots)),
        ("Earliest snapshot",        _snap_date(snapshots[0]) if snapshots else "—"),
        ("Latest snapshot",          _snap_date(snapshots[-1]) if snapshots else "—"),
        ("Projects in latest snap",  latest.get("projects_in_scope", "—")),
        ("Succeeded in latest snap", latest.get("projects_success", "—")),
        ("Failed in latest snap",    latest.get("projects_failed",  "—")),
        ("Skipped in latest snap",   latest.get("projects_skipped", "—")),
        ("Dept mapping entries",     len(dept_mapping)),
    ]
    for row_idx, (metric, value) in enumerate(rows, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        _cell(ws, row_idx, 1, metric, fill=fill, bold=True)
        _cell(ws, row_idx, 2, value,  fill=fill)

    _auto_width(ws)


def _build_fleet_trend_sheet(ws, fleet: dict) -> None:
    ws.title = "Fleet Trend"

    dates   = fleet["dates"]
    n_dates = len(dates)
    if n_dates == 0:
        _cell(ws, 1, 1, "No snapshot data available.")
        return

    # ── Data table ─────────────────────────────────────────────────────────────
    # Row 1: header row — "Metric" then dates
    _hdr(ws, 1, 1, "Metric")
    for col_idx, d in enumerate(dates, 2):
        _hdr(ws, 1, col_idx, d)
    ws.row_dimensions[1].height = 28

    row_labels = [
        ("Active Vulns",       fleet["active"]),
        ("Acknowledged Vulns", fleet["acknowledged"]),
        ("To Review Hotspots", fleet["to_review"]),
    ]

    for row_idx, (label, values) in enumerate(row_labels, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        _cell(ws, row_idx, 1, label, fill=fill, bold=True)
        for col_idx, val in enumerate(values, 2):
            _cell(ws, row_idx, col_idx, val, fill=fill, align=_RIGHT)

    # Net change row (active delta)
    net_row = len(row_labels) + 2
    _cell(ws, net_row, 1, "Net Change (Active Δ)", fill=_WHITE_FILL, bold=True)
    active = fleet["active"]
    for col_idx in range(2, n_dates + 2):
        idx = col_idx - 2
        if idx == 0:
            val = 0
        else:
            val = active[idx] - active[idx - 1]
        fill = _RED_FILL if (isinstance(val, int) and val > 0) else _GREEN_FILL
        _cell(ws, net_row, col_idx, val, fill=fill, align=_RIGHT)

    # ── Auto width ─────────────────────────────────────────────────────────────
    _auto_width(ws)

    # ── Chart: Active Vulns + Acknowledged ────────────────────────────────────
    # Data rows 2 and 3 (active + acknowledged); categories = row 1 from col 2
    data_ref = Reference(ws, min_row=2, max_row=3,
                         min_col=2, max_col=n_dates + 1)
    cats_ref = Reference(ws, min_row=1, min_col=2, max_col=n_dates + 1)
    anchor   = f"A{net_row + 2}"
    _add_line_chart(ws, "Fleet Vulnerability Trend", anchor, data_ref, cats_ref)


def _build_dept_sheet(ws, dept_name: str, dept_data: dict,
                      dept_mapping: dict[str, dict]) -> None:
    ws.title = f"Dept - {dept_name}"[:31]

    dates   = dept_data["dates"]
    codes   = dept_data["codes"]   # {code: [values]}
    totals  = dept_data["totals"]
    n_dates = len(dates)

    if n_dates == 0 or not codes:
        _cell(ws, 1, 1, "No data available.")
        return

    sorted_codes = sorted(codes.keys())

    # ── Header row ─────────────────────────────────────────────────────────────
    _hdr(ws, 1, 1, "Code")
    for col_idx, d in enumerate(dates, 2):
        _hdr(ws, 1, col_idx, d)
    ws.row_dimensions[1].height = 28

    # ── Data rows: one per CODE ────────────────────────────────────────────────
    for row_idx, code in enumerate(sorted_codes, 2):
        entry = dept_mapping.get(code, {})
        internet_facing = entry.get("internet_facing", False)
        critical        = entry.get("critical", False)
        if critical:
            fill = _RED_FILL
        elif internet_facing:
            fill = _YELLOW_FILL
        else:
            fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL

        _cell(ws, row_idx, 1, code, fill=fill, bold=True)
        for col_idx, val in enumerate(codes[code], 2):
            _cell(ws, row_idx, col_idx, val, fill=fill, align=_RIGHT)

    # ── Dept total row ─────────────────────────────────────────────────────────
    total_row = len(sorted_codes) + 2
    _cell(ws, total_row, 1, f"TOTAL ({dept_name})", fill=_TOTAL_FILL, bold=True)
    ws.cell(row=total_row, column=1).font = _TOTAL_FONT
    for col_idx, val in enumerate(totals, 2):
        _cell(ws, total_row, col_idx, val, fill=_TOTAL_FILL, align=_RIGHT, bold=True)
        ws.cell(row=total_row, column=col_idx).font = _TOTAL_FONT

    _auto_width(ws)

    # ── Chart: top MAX_CHART_SERIES codes by latest snapshot value ─────────────
    if n_dates >= 1 and len(sorted_codes) >= 1:
        # Determine which code rows to chart
        latest_idx = n_dates - 1
        ranked = sorted(sorted_codes,
                        key=lambda c: codes[c][latest_idx] if codes[c] else 0,
                        reverse=True)
        chart_codes = ranked[:MAX_CHART_SERIES]
        chart_rows  = [sorted_codes.index(c) + 2 for c in chart_codes]

        if chart_rows:
            chart = LineChart()
            chart.title  = f"Dept: {dept_name} — Active Vulns by Code"
            chart.style  = 10
            chart.width  = CHART_WIDTH
            chart.height = CHART_HEIGHT
            chart.y_axis.title = "Active Vulnerabilities"
            chart.x_axis.title = "Date"

            for row_idx in chart_rows:
                data_ref = Reference(ws, min_row=row_idx, max_row=row_idx,
                                     min_col=2, max_col=n_dates + 1)
                from openpyxl.chart import Series
                series = Series(data_ref,
                                title=Reference(ws, min_row=row_idx, max_row=row_idx,
                                                min_col=1, max_col=1))
                chart.series.append(series)

            cats_ref = Reference(ws, min_row=1, min_col=2, max_col=n_dates + 1)
            chart.set_categories(cats_ref)
            anchor = f"A{total_row + 2}"
            ws.add_chart(chart, anchor)


def _build_code_sheet(ws, code_name: str, code_data: dict,
                      dept_mapping: dict[str, dict]) -> None:
    ws.title = f"Code - {code_name}"[:31]

    dates    = code_data["dates"]
    projects = code_data["projects"]  # {key: [values]}
    n_dates  = len(dates)

    if n_dates == 0 or not projects:
        _cell(ws, 1, 1, "No data available.")
        return

    entry           = dept_mapping.get(code_name, {})
    internet_facing = entry.get("internet_facing", False)
    critical        = entry.get("critical", False)
    dept            = entry.get("dept", "Unmapped")

    sorted_proj_keys = sorted(projects.keys())

    # ── Header: fixed cols then dates ─────────────────────────────────────────
    _hdr(ws, 1, 1, "Project Key")
    _hdr(ws, 1, 2, "Dept")
    _hdr(ws, 1, 3, "Internet Facing")
    _hdr(ws, 1, 4, "Critical")
    date_col_start = 5
    for col_idx, d in enumerate(dates, date_col_start):
        _hdr(ws, 1, col_idx, d)
    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, proj_key in enumerate(sorted_proj_keys, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        _cell(ws, row_idx, 1, proj_key,                              fill=fill)
        _cell(ws, row_idx, 2, dept,                                  fill=fill)
        _cell(ws, row_idx, 3, "Yes" if internet_facing else "No",    fill=fill, align=_CENTER)
        _cell(ws, row_idx, 4, "Yes" if critical else "No",           fill=fill, align=_CENTER)
        for col_idx, val in enumerate(projects[proj_key], date_col_start):
            _cell(ws, row_idx, col_idx, val, fill=fill, align=_RIGHT)

    _auto_width(ws)

    # ── Chart: top MAX_CHART_SERIES projects by latest value ──────────────────
    n_proj = len(sorted_proj_keys)
    if n_dates >= 1 and n_proj >= 1:
        latest_idx = n_dates - 1
        ranked = sorted(sorted_proj_keys,
                        key=lambda k: projects[k][latest_idx] if projects[k] else 0,
                        reverse=True)
        chart_keys = ranked[:MAX_CHART_SERIES]
        chart_rows = [sorted_proj_keys.index(k) + 2 for k in chart_keys]
        max_data_row = len(sorted_proj_keys) + 2

        if chart_rows:
            chart = LineChart()
            chart.title  = f"Code: {code_name} — Active Vulns by Project"
            chart.style  = 10
            chart.width  = CHART_WIDTH
            chart.height = CHART_HEIGHT
            chart.y_axis.title = "Active Vulnerabilities"
            chart.x_axis.title = "Date"

            for row_idx in chart_rows:
                data_ref = Reference(ws, min_row=row_idx, max_row=row_idx,
                                     min_col=date_col_start,
                                     max_col=date_col_start + n_dates - 1)
                from openpyxl.chart import Series
                series = Series(data_ref,
                                title=Reference(ws, min_row=row_idx, max_row=row_idx,
                                                min_col=1, max_col=1))
                chart.series.append(series)

            cats_ref = Reference(ws, min_row=1,
                                 min_col=date_col_start,
                                 max_col=date_col_start + n_dates - 1)
            chart.set_categories(cats_ref)
            anchor = f"A{max_data_row + 2}"
            ws.add_chart(chart, anchor)


def _build_events_sheet(ws, events: list[dict]) -> None:
    ws.title = "Events"
    ws.freeze_panes = "A2"

    headers = ["Date", "Event Type", "Project Key", "Project Name",
               "Code", "Dept", "Detail"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    def _event_fill(event_type: str) -> PatternFill:
        return {
            "Branch Changed":   _BLUE_FILL,
            "Project Added":    _GREEN_FILL,
            "Project Removed":  _ORANGE_FILL,
            "Unmapped Code":    _YELLOW_FILL,
            "Branch Stale":     _LORANGE_FILL,
            "Snapshot Partial": _LORANGE_FILL,
        }.get(event_type, _WHITE_FILL)

    for row_idx, ev in enumerate(events, 2):
        fill = _event_fill(ev["event_type"])
        _cell(ws, row_idx, 1, ev["date"],         fill=fill, align=_CENTER)
        _cell(ws, row_idx, 2, ev["event_type"],   fill=fill, align=_CENTER)
        _cell(ws, row_idx, 3, ev["project_key"],  fill=fill)
        _cell(ws, row_idx, 4, ev["project_name"], fill=fill)
        _cell(ws, row_idx, 5, ev["code"],         fill=fill)
        _cell(ws, row_idx, 6, ev["dept"],         fill=fill)
        _cell(ws, row_idx, 7, ev["detail"],       fill=fill)

    if not events:
        _cell(ws, 2, 1, "No events detected.", fill=_WHITE_FILL)

    _auto_width(ws)


def _build_coverage_gaps_sheet(
    ws,
    snapshots: list[dict],
    dept_mapping: dict[str, dict],
) -> None:
    ws.title = "Coverage Gaps"

    if not snapshots:
        _cell(ws, 1, 1, "No snapshot data available.")
        return

    latest = snapshots[-1]

    # ── Section 1: Inactive / skipped projects ─────────────────────────────────
    row = 1
    _hdr(ws, row, 1, "Section 1: Inactive Projects (no valid branch in latest snapshot)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 24
    row += 1

    sub_hdrs = ["Project Key", "Project Name", "Code", "Dept", "Internet Facing", "Critical"]
    for col, h in enumerate(sub_hdrs, 1):
        _hdr(ws, row, col, h)
    ws.row_dimensions[row].height = 26
    row += 1

    inactive_start = row
    for proj in latest.get("projects", []):
        err = proj.get("error", "")
        if err and err.startswith("skipped"):
            key  = proj.get("key", "")
            name = proj.get("name", key)
            code = proj.get("code") or _extract_code(name)
            entry = dept_mapping.get(code, {})
            fill = _ALT_FILL if row % 2 == 0 else _WHITE_FILL
            _cell(ws, row, 1, key,                                              fill=fill)
            _cell(ws, row, 2, name,                                             fill=fill)
            _cell(ws, row, 3, code,                                             fill=fill)
            _cell(ws, row, 4, entry.get("dept", "Unmapped"),                   fill=fill)
            _cell(ws, row, 5, "Yes" if entry.get("internet_facing") else "No", fill=fill, align=_CENTER)
            _cell(ws, row, 6, "Yes" if entry.get("critical") else "No",        fill=fill, align=_CENTER)
            row += 1

    if row == inactive_start:
        _cell(ws, row, 1, "No inactive projects.", fill=_WHITE_FILL)
        row += 1

    row += 1  # blank separator

    # ── Section 2: Unmapped codes ──────────────────────────────────────────────
    _hdr(ws, row, 1, "Section 2: Unmapped Codes (codes not in dept mapping)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 24
    row += 1

    sub_hdrs2 = ["Code", "Example Project", "Count of Projects"]
    for col, h in enumerate(sub_hdrs2, 1):
        _hdr(ws, row, col, h)
    ws.row_dimensions[row].height = 26
    row += 1

    # Collect unmapped codes from latest snapshot
    unmapped: dict[str, dict] = {}  # code → {example, count}
    for proj in latest.get("projects", []):
        key  = proj.get("key", "")
        name = proj.get("name", key)
        code = proj.get("code") or _extract_code(name)
        if code not in dept_mapping:
            if code not in unmapped:
                unmapped[code] = {"example": name, "count": 0}
            unmapped[code]["count"] += 1

    if unmapped:
        for fill_row_idx, (code, info) in enumerate(sorted(unmapped.items()), 0):
            fill = _ALT_FILL if fill_row_idx % 2 == 0 else _WHITE_FILL
            _cell(ws, row, 1, code,            fill=fill)
            _cell(ws, row, 2, info["example"], fill=fill)
            _cell(ws, row, 3, info["count"],   fill=fill, align=_RIGHT)
            row += 1
    else:
        _cell(ws, row, 1, "No unmapped codes.", fill=_WHITE_FILL)

    _auto_width(ws)


# ── Main report builder ────────────────────────────────────────────────────────

def build_report(args: argparse.Namespace) -> None:
    # ── Load data ──────────────────────────────────────────────────────────────
    snapshots = load_snapshots(args.snapshots_dir, args.snapshots_count)
    if not snapshots:
        print(f"No snapshots found in {args.snapshots_dir}.", file=sys.stderr)
        sys.exit(1)

    tqdm.write(f"\nLoaded {len(snapshots)} snapshot(s).")

    dept_mapping = load_dept_mapping(args.mapping)
    tqdm.write(f"Loaded {len(dept_mapping)} dept mapping entries.\n")

    # ── Optional project filter ────────────────────────────────────────────────
    filter_keys: Optional[set[str]] = None
    if args.projects:
        filter_keys = set(load_project_keys_file(args.projects))
        tqdm.write(f"Filtering to {len(filter_keys)} project keys.")

    # ── Apply filters ──────────────────────────────────────────────────────────
    snapshots = _filter_projects_in_snapshots(
        snapshots, filter_keys, dept_mapping,
        args.internet_facing_only, args.critical_only,
    )

    # ── Build time series ──────────────────────────────────────────────────────
    tqdm.write("Building time series...")
    fleet_data  = build_fleet_series(snapshots)
    dept_data   = build_dept_series(snapshots, dept_mapping)
    code_data   = build_code_series(snapshots)

    # ── Detect events ──────────────────────────────────────────────────────────
    events = detect_events(snapshots, dept_mapping)
    tqdm.write(f"Detected {len(events)} event(s).")

    # ── Build Excel ────────────────────────────────────────────────────────────
    tqdm.write("Writing Excel report...")
    wb = Workbook()

    # Overview sheet (first)
    _build_overview_sheet(wb.active, snapshots, dept_mapping)

    # Fleet trend
    _build_fleet_trend_sheet(wb.create_sheet(), fleet_data)

    # Dept sheets (sorted alphabetically)
    for dept_name in sorted(dept_data.keys()):
        _build_dept_sheet(wb.create_sheet(), dept_name, dept_data[dept_name], dept_mapping)

    # Code sheets (sorted alphabetically)
    for code_name in tqdm(sorted(code_data.keys()), desc="Writing code sheets", unit="sheet"):
        _build_code_sheet(wb.create_sheet(), code_name, code_data[code_name], dept_mapping)

    # Events
    _build_events_sheet(wb.create_sheet(), events)

    # Coverage Gaps
    _build_coverage_gaps_sheet(wb.create_sheet(), snapshots, dept_mapping)

    wb.save(args.output)
    tqdm.write(f"\nExcel report saved → {args.output}")

    print("\n" + "═" * 60)
    print("  TREND REPORT SUMMARY")
    print("═" * 60)
    print(f"  Snapshots loaded     : {len(snapshots)}")
    print(f"  Dept sheets          : {len(dept_data)}")
    print(f"  Code sheets          : {len(code_data)}")
    print(f"  Events detected      : {len(events)}")
    print(f"  Output               : {args.output}")
    print("═" * 60)


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate a multi-sheet Excel trend report from SonarQube snapshots.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--snapshots-dir", default=os.getenv("SONAR_SNAPSHOTS_DIR", "snapshots"),
                   help="Directory containing snapshot JSON files  [env: SONAR_SNAPSHOTS_DIR]")
    p.add_argument("--mapping", default=os.getenv("SONAR_DEPT_MAPPING"),
                   help="Path to department mapping CSV  [env: SONAR_DEPT_MAPPING]")
    p.add_argument("--output",  default=os.getenv("SONAR_TREND_OUTPUT", "trend_report.xlsx"),
                   help="Path for the Excel output file  [env: SONAR_TREND_OUTPUT]")
    p.add_argument("--projects", default=os.getenv("SONAR_PROJECTS_FILE"),
                   help="Optional text file with project keys to restrict the report  "
                        "[env: SONAR_PROJECTS_FILE]")
    p.add_argument("--snapshots-count", type=int,
                   default=int(os.getenv("SONAR_SNAPSHOTS_COUNT", DEFAULT_SNAPSHOTS_COUNT)),
                   help="Number of most-recent snapshots to load  [env: SONAR_SNAPSHOTS_COUNT]")
    p.add_argument("--internet-facing-only", action="store_true",
                   help="Only include projects whose CODE is marked internetFacing=true "
                        "in the mapping file")
    p.add_argument("--critical-only", action="store_true",
                   help="Only include projects whose CODE is marked Critical=true "
                        "in the mapping file")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    if not args.mapping:
        print("Error: --mapping is required (or set SONAR_DEPT_MAPPING).",
              file=sys.stderr)
        sys.exit(1)

    print("SonarQube Trend Report")
    print(f"  Snapshots dir    : {args.snapshots_dir}")
    print(f"  Snapshots count  : {args.snapshots_count}")
    print(f"  Mapping file     : {args.mapping}")
    print(f"  Output           : {args.output}")
    print(f"  Internet facing  : {args.internet_facing_only}")
    print(f"  Critical only    : {args.critical_only}")
    print()

    build_report(args)


if __name__ == "__main__":
    main()
