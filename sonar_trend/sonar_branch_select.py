#!/usr/bin/env python3
"""
SonarQube Branch Selector
=========================
Evaluates ALL SonarQube projects and selects the best representative branch per
project.  Saves branch_selection.json and an Excel report.

Branch selection priority per project:
  1. isMain branch — only if LOC > 0 (checked via /api/measures/component)
  2. Named fallbacks in order: develop, development, dev, master — first with LOC > 0
  3. Any other scanned branch (has analysisDate) with LOC > 0, most recently scanned first
  4. None — no valid branch found → project marked as "Inactive"

A selected branch is "stale" if its analysisDate is > 60 days ago.

If a previous branch_selection.json exists at the output path, the script
detects branch changes and records previous_branch / branch_changed_at.

Usage:
    python sonar_branch_select.py \\
        --url http://sonarqube.example.com \\
        --username admin \\
        --password secret

    # Custom output, higher concurrency
    python sonar_branch_select.py \\
        --url http://sonar.internal \\
        --username admin --password secret \\
        --output branch_selection.json \\
        --excel  branch_selection_report.xlsx \\
        --concurrency 20

    # Skip SSL verification (self-signed certs)
    python sonar_branch_select.py ... --no-verify-ssl

Requirements:
    pip install httpx openpyxl tqdm python-dotenv
"""

import asyncio
import httpx
import json
import argparse
import sys
import os
from typing import Any, Optional
from datetime import datetime, timezone, timedelta

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from tqdm import tqdm
from tqdm.asyncio import tqdm as atqdm

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ──────────────────────────────────────────────────────────────────

PAGE_SIZE           = 500
DEFAULT_CONCURRENCY = 10
STALE_THRESHOLD_DAYS = 60

FALLBACK_NAMES = ["develop", "development", "dev", "master"]

REASON_ISMAIN_VALID    = "isMain-valid"
REASON_ISMAIN_STALE    = "isMain-stale"
REASON_FALLBACK_PREFIX = "fallback-"
REASON_ANY_RECENT      = "any-recent"
REASON_ANY_STALE       = "any-stale"
REASON_NONE            = "none"


# ── SonarQube Client ───────────────────────────────────────────────────────────

class SonarQubeClient:
    def __init__(self, base_url: str, username: str, password: str,
                 concurrency: int = DEFAULT_CONCURRENCY,
                 verify_ssl: bool = True):
        self.base_url  = base_url.rstrip("/")
        self.auth      = (username, password)
        self.verify_ssl = verify_ssl
        self._sem      = asyncio.Semaphore(concurrency)

    async def _get(self, client: httpx.AsyncClient, path: str,
                   params: dict[str, Any] | None = None) -> dict:
        async with self._sem:
            resp = await client.get(
                f"{self.base_url}{path}",
                params=params or {},
                auth=self.auth,
                timeout=60.0,
            )
            resp.raise_for_status()
            return resp.json()

    async def get_all_projects(self, client: httpx.AsyncClient) -> list[dict]:
        """Fetch all projects using pagination with a tqdm progress bar."""
        projects: list[dict] = []
        page = 1
        pbar = tqdm(desc="Fetching projects", unit="project", total=None)
        try:
            while True:
                data = await self._get(client, "/api/projects/search",
                                       {"ps": PAGE_SIZE, "p": page})
                batch = data.get("components", [])
                projects.extend(batch)
                pg = data.get("paging", {})
                if page == 1:
                    pbar.total = pg.get("total", 0)
                    pbar.refresh()
                pbar.update(len(batch))
                if pg.get("pageIndex", 1) * pg.get("pageSize", PAGE_SIZE) >= pg.get("total", 0):
                    break
                page += 1
        finally:
            pbar.close()
        return projects

    async def get_branches(self, client: httpx.AsyncClient,
                           project_key: str) -> list[dict]:
        data = await self._get(client, "/api/project_branches/list",
                               {"project": project_key})
        return data.get("branches", [])

    async def get_ncloc(self, client: httpx.AsyncClient,
                        project_key: str, branch_name: str) -> int:
        try:
            data = await self._get(client, "/api/measures/component", {
                "component": project_key,
                "branch": branch_name,
                "metricKeys": "ncloc",
            })
            measures = data.get("component", {}).get("measures", [])
            if measures:
                return int(measures[0].get("value", 0))
            return 0
        except Exception:
            return 0


# ── Helpers ────────────────────────────────────────────────────────────────────

def _warn(msg: str) -> None:
    tqdm.write(f"  [WARN] {msg}", file=sys.stderr)


def _extract_code(name: str) -> str:
    return name.split("_")[0] if "_" in name else name


def _parse_iso(iso: Optional[str]) -> Optional[datetime]:
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except ValueError:
        return None


def _is_stale(analysis_date_iso: Optional[str], threshold_days: int = STALE_THRESHOLD_DAYS) -> bool:
    dt = _parse_iso(analysis_date_iso)
    if not dt:
        return False
    now = datetime.now(timezone.utc)
    return (now - dt) > timedelta(days=threshold_days)


def load_previous_selection(path: str) -> dict:
    """Load existing branch_selection.json; return empty dict if missing."""
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
            return data.get("projects", {})
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


# ── Branch Selection Logic ─────────────────────────────────────────────────────

async def select_branch_for_project(
    sonar: SonarQubeClient,
    client: httpx.AsyncClient,
    project_key: str,
    project_name: str,
    branches: list[dict],
    previous_projects: dict,
    run_ts: str,
    pbar: tqdm,
) -> dict:
    """Evaluate branches and pick the best one.  Returns a project result dict."""
    try:
        result = await _do_select(sonar, client, project_key, project_name,
                                  branches, previous_projects, run_ts)
    except Exception as exc:
        _warn(f"Failed to select branch for {project_key}: {exc}")
        result = {
            "name": project_name,
            "code": _extract_code(project_name),
            "branch": None,
            "reason": REASON_NONE,
            "ncloc": 0,
            "analysis_date": None,
            "is_stale": False,
            "previous_branch": None,
            "branch_changed_at": None,
        }
    finally:
        pbar.update(1)
    return result


async def _do_select(
    sonar: SonarQubeClient,
    client: httpx.AsyncClient,
    project_key: str,
    project_name: str,
    branches: list[dict],
    previous_projects: dict,
    run_ts: str,
) -> dict:
    code = _extract_code(project_name)
    prev = previous_projects.get(project_key, {})
    prev_branch = prev.get("branch")

    # Build lookup by branch name
    branch_by_name: dict[str, dict] = {b["name"]: b for b in branches}

    selected_branch: Optional[str] = None
    selected_ncloc: int = 0
    selected_analysis_date: Optional[str] = None
    reason: str = REASON_NONE

    # ── Step 1: isMain branch ──────────────────────────────────────────────────
    main_branches = [b for b in branches if b.get("isMain", False)]
    if main_branches:
        main_b = main_branches[0]
        ncloc = await sonar.get_ncloc(client, project_key, main_b["name"])
        if ncloc > 0:
            selected_branch = main_b["name"]
            selected_ncloc = ncloc
            selected_analysis_date = main_b.get("analysisDate")
            stale = _is_stale(selected_analysis_date)
            reason = REASON_ISMAIN_STALE if stale else REASON_ISMAIN_VALID

    # ── Step 2: named fallbacks ────────────────────────────────────────────────
    if selected_branch is None:
        for fb_name in FALLBACK_NAMES:
            if fb_name in branch_by_name:
                b = branch_by_name[fb_name]
                ncloc = await sonar.get_ncloc(client, project_key, fb_name)
                if ncloc > 0:
                    selected_branch = fb_name
                    selected_ncloc = ncloc
                    selected_analysis_date = b.get("analysisDate")
                    reason = REASON_FALLBACK_PREFIX + fb_name
                    break

    # ── Step 3: any scanned branch, most recent first ──────────────────────────
    if selected_branch is None:
        scanned = [b for b in branches if b.get("analysisDate")]
        scanned.sort(key=lambda b: b.get("analysisDate", ""), reverse=True)
        for b in scanned:
            ncloc = await sonar.get_ncloc(client, project_key, b["name"])
            if ncloc > 0:
                selected_branch = b["name"]
                selected_ncloc = ncloc
                selected_analysis_date = b.get("analysisDate")
                stale = _is_stale(selected_analysis_date)
                reason = REASON_ANY_STALE if stale else REASON_ANY_RECENT
                break

    # ── Staleness override for fallback/any-recent ─────────────────────────────
    is_stale_flag = _is_stale(selected_analysis_date) if selected_branch else False
    if selected_branch and reason == REASON_ANY_RECENT and is_stale_flag:
        reason = REASON_ANY_STALE

    # ── Branch change detection ────────────────────────────────────────────────
    branch_changed_at = prev.get("branch_changed_at")
    if prev_branch is not None and selected_branch != prev_branch:
        branch_changed_at = run_ts
    elif prev_branch is None:
        # first run — no previous
        branch_changed_at = None

    return {
        "name": project_name,
        "code": code,
        "branch": selected_branch,
        "reason": reason,
        "ncloc": selected_ncloc,
        "analysis_date": selected_analysis_date,
        "is_stale": is_stale_flag,
        "previous_branch": prev_branch,
        "branch_changed_at": branch_changed_at,
    }


# ── Orchestration ──────────────────────────────────────────────────────────────

async def run(args: argparse.Namespace) -> dict:
    sonar = SonarQubeClient(
        base_url=args.url,
        username=args.username,
        password=args.password,
        concurrency=args.concurrency,
        verify_ssl=not args.no_verify_ssl,
    )

    run_ts = datetime.now(timezone.utc).isoformat()
    previous_projects = load_previous_selection(args.output)
    if previous_projects:
        tqdm.write(f"  Loaded {len(previous_projects)} entries from previous selection.")

    async with httpx.AsyncClient(verify=not args.no_verify_ssl) as client:

        # ── Phase 1: fetch all projects ────────────────────────────────────────
        projects = await sonar.get_all_projects(client)
        tqdm.write(f"\nFound {len(projects)} project(s).\n")

        # ── Phase 2: fetch branch lists concurrently ───────────────────────────
        async def fetch_branches(p: dict) -> list[dict]:
            try:
                return await sonar.get_branches(client, p["key"])
            except Exception as exc:
                _warn(f"Could not fetch branches for {p['key']}: {exc}")
                return []

        branches_list: list[list[dict]] = await atqdm.gather(
            *[fetch_branches(p) for p in projects],
            desc="Fetching branch lists",
            unit="project",
            total=len(projects),
        )

        tqdm.write("")

        # ── Phase 3: select branch per project concurrently ───────────────────
        with tqdm(desc="Selecting branches", unit="project",
                  total=len(projects)) as pbar:
            results: list[dict] = list(await asyncio.gather(*[
                select_branch_for_project(
                    sonar, client,
                    projects[i]["key"],
                    projects[i]["name"],
                    branches_list[i],
                    previous_projects,
                    run_ts,
                    pbar,
                )
                for i in range(len(projects))
            ]))

    # ── Build output structure ─────────────────────────────────────────────────
    projects_dict: dict[str, dict] = {}
    for i, proj in enumerate(projects):
        projects_dict[proj["key"]] = results[i]

    valid    = sum(1 for r in results if r["branch"] and r["reason"] in (REASON_ISMAIN_VALID,))
    fallback = sum(1 for r in results if r["branch"] and r["reason"].startswith(REASON_FALLBACK_PREFIX))
    stale    = sum(1 for r in results if r["branch"] and r["is_stale"])
    inactive = sum(1 for r in results if not r["branch"])

    output = {
        "generated_at": run_ts,
        "sonar_url": args.url,
        "stale_threshold_days": STALE_THRESHOLD_DAYS,
        "total_projects": len(projects),
        "valid": valid,
        "fallback": fallback,
        "stale": stale,
        "inactive": inactive,
        "projects": projects_dict,
    }
    return output


# ── Console summary ────────────────────────────────────────────────────────────

def print_summary(data: dict) -> None:
    print("\n" + "═" * 60)
    print("  BRANCH SELECTION SUMMARY")
    print("═" * 60)
    print(f"  Total projects   : {data['total_projects']}")
    print(f"  Valid (isMain)   : {data['valid']}")
    print(f"  Fallback         : {data['fallback']}")
    print(f"  Stale            : {data['stale']}")
    print(f"  Inactive (none)  : {data['inactive']}")
    print("═" * 60)


# ── Excel styles ───────────────────────────────────────────────────────────────

_HDR_FILL      = PatternFill("solid", fgColor="1F3864")
_TOTAL_FILL    = PatternFill("solid", fgColor="375623")
_ALT_FILL      = PatternFill("solid", fgColor="DEEAF1")
_WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
_GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")   # isMain-valid
_YELLOW_FILL   = PatternFill("solid", fgColor="FFFF99")   # fallback
_ORANGE_FILL   = PatternFill("solid", fgColor="FCE4D6")   # any-stale

_HDR_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT     = Font(name="Calibri", size=10)
_BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
_TOTAL_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

_THIN          = Side(border_style="thin", color="B8CCE4")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center")
_RIGHT         = Alignment(horizontal="right",  vertical="center")


def _hdr(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _HDR_FILL; c.font = _HDR_FONT
    c.alignment = _CENTER; c.border = _BORDER


def _cell(ws, row: int, col: int, value,
          fill=None, align=None, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill or _WHITE_FILL
    c.font      = _BOLD_FONT if bold else _BODY_FONT
    c.alignment = align or _LEFT
    c.border    = _BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _row_fill_for_reason(reason: str) -> PatternFill:
    if reason == REASON_ISMAIN_VALID:
        return _GREEN_FILL
    if reason == REASON_ANY_STALE or reason == REASON_ISMAIN_STALE:
        return _ORANGE_FILL
    if reason.startswith(REASON_FALLBACK_PREFIX) or reason == REASON_ANY_RECENT:
        return _YELLOW_FILL
    return _WHITE_FILL


# ── Excel sheet builders ───────────────────────────────────────────────────────

def _build_summary_sheet(ws, data: dict) -> None:
    ws.title = "Summary"
    ws.freeze_panes = "A2"

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 28

    rows = [
        ("Generated at",       data["generated_at"]),
        ("SonarQube URL",      data["sonar_url"]),
        ("Stale threshold (days)", data["stale_threshold_days"]),
        ("Total projects",     data["total_projects"]),
        ("Valid (isMain)",     data["valid"]),
        ("Fallback branches",  data["fallback"]),
        ("Stale branches",     data["stale"]),
        ("Inactive projects",  data["inactive"]),
    ]
    for row_idx, (metric, value) in enumerate(rows, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        _cell(ws, row_idx, 1, metric, fill=fill, bold=True)
        _cell(ws, row_idx, 2, value,  fill=fill)

    _auto_width(ws)


def _build_valid_sheet(ws, data: dict) -> None:
    ws.title = "Valid"
    ws.freeze_panes = "A2"

    headers = [
        "Project Key", "Project Name", "Code", "Department",
        "Branch", "Reason", "LOC", "Analysis Date",
        "Is Stale", "Previous Branch", "Branch Changed At",
    ]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    valid_projects = [
        (key, proj) for key, proj in data["projects"].items()
        if proj["branch"] is not None
    ]
    valid_projects.sort(key=lambda x: x[1]["name"].lower())

    for row_idx, (key, proj) in enumerate(valid_projects, 2):
        fill = _row_fill_for_reason(proj["reason"])
        _cell(ws, row_idx,  1, key,                          fill=fill)
        _cell(ws, row_idx,  2, proj["name"],                 fill=fill)
        _cell(ws, row_idx,  3, proj["code"],                 fill=fill)
        _cell(ws, row_idx,  4, "",                           fill=fill)  # dept unknown here
        _cell(ws, row_idx,  5, proj["branch"],               fill=fill)
        _cell(ws, row_idx,  6, proj["reason"],               fill=fill, align=_CENTER)
        _cell(ws, row_idx,  7, proj["ncloc"],                fill=fill, align=_RIGHT)
        _cell(ws, row_idx,  8, proj["analysis_date"] or "",  fill=fill, align=_CENTER)
        _cell(ws, row_idx,  9, "Yes" if proj["is_stale"] else "No",
                                                              fill=fill, align=_CENTER)
        _cell(ws, row_idx, 10, proj["previous_branch"] or "", fill=fill)
        _cell(ws, row_idx, 11, proj["branch_changed_at"] or "", fill=fill)

    _auto_width(ws)


def _build_inactive_sheet(ws, data: dict) -> None:
    ws.title = "Inactive"
    ws.freeze_panes = "A2"

    headers = ["Project Key", "Project Name", "Code"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    inactive = [
        (key, proj) for key, proj in data["projects"].items()
        if proj["branch"] is None
    ]
    inactive.sort(key=lambda x: x[1]["name"].lower())

    for row_idx, (key, proj) in enumerate(inactive, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        _cell(ws, row_idx, 1, key,         fill=fill)
        _cell(ws, row_idx, 2, proj["name"], fill=fill)
        _cell(ws, row_idx, 3, proj["code"], fill=fill)

    total_row = len(inactive) + 2
    _cell(ws, total_row, 1, f"TOTAL  ({len(inactive)} project(s))",
          fill=_TOTAL_FILL, bold=True)
    ws.cell(row=total_row, column=1).font = _TOTAL_FONT
    for col in [2, 3]:
        ws.cell(row=total_row, column=col).fill   = _TOTAL_FILL
        ws.cell(row=total_row, column=col).border = _BORDER

    _auto_width(ws)


def export_excel(data: dict, path: str) -> None:
    wb = Workbook()
    _build_summary_sheet(wb.active, data)
    _build_valid_sheet(wb.create_sheet(), data)
    _build_inactive_sheet(wb.create_sheet(), data)
    wb.save(path)


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Select the best representative branch per SonarQube project.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--url",      default=os.getenv("SONAR_URL"),
                   help="SonarQube base URL  [env: SONAR_URL]")
    p.add_argument("--username", default=os.getenv("SONAR_USERNAME"),
                   help="SonarQube username  [env: SONAR_USERNAME]")
    p.add_argument("--password", default=os.getenv("SONAR_PASSWORD"),
                   help="SonarQube password  [env: SONAR_PASSWORD]")
    p.add_argument("--output",   default=os.getenv("SONAR_BRANCH_CACHE", "branch_selection.json"),
                   help="Path for the JSON output file  [env: SONAR_BRANCH_CACHE]")
    p.add_argument("--excel",    default=os.getenv("SONAR_BRANCH_EXCEL", "branch_selection_report.xlsx"),
                   help="Path for the Excel output file  [env: SONAR_BRANCH_EXCEL]")
    p.add_argument("--concurrency", type=int,
                   default=int(os.getenv("SONAR_CONCURRENCY", DEFAULT_CONCURRENCY)),
                   help="Max simultaneous API requests  [env: SONAR_CONCURRENCY]")
    p.add_argument("--no-verify-ssl", action="store_true",
                   help="Disable SSL certificate verification (self-signed certs)")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    missing = [f"--{k}" for k, v in
               [("url", args.url), ("username", args.username),
                ("password", args.password)] if not v]
    if missing:
        print(f"Error: missing required values: {', '.join(missing)}\n"
              f"Pass them as CLI args or set SONAR_URL / SONAR_USERNAME / SONAR_PASSWORD "
              f"in your environment / .env file.",
              file=sys.stderr)
        sys.exit(1)

    print("SonarQube Branch Selector")
    print(f"  URL         : {args.url}")
    print(f"  Username    : {args.username}")
    print(f"  Concurrency : {args.concurrency}")
    print(f"  Output      : {args.output}")
    print(f"  Excel       : {args.excel}")
    print(f"  Verify SSL  : {not args.no_verify_ssl}")
    print()

    data = asyncio.run(run(args))

    print_summary(data)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"\nJSON saved  → {args.output}")

    export_excel(data, args.excel)
    print(f"Excel saved → {args.excel}")


if __name__ == "__main__":
    main()
