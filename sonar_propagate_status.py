#!/usr/bin/env python3
"""
SonarQube Status Propagator
============================
Propagates acknowledged issue statuses from a source branch to matching OPEN
issues on all other active branches of the same project(s).

Statuses propagated from source → transition applied on target:
  CONFIRMED       → confirm
  FALSE_POSITIVE  → falsepositive
  ACCEPTED        → accept

Matching is done by (rule, component, hash) — the same logical issue across
branches.  Target issues that already carry any non-OPEN status are skipped
(their status was a deliberate choice on that branch).

Branches with no analysisDate or last scanned > 60 days ago are skipped.

Input:
  --project <key>      a single project key
  --projects <file>    a text file with one project key per line
  (both can be combined; keys are merged and deduplicated)

Output:
  - Console progress and summary
  - Excel report: Summary sheet (per project+branch) + Detail sheet (per issue)

Usage:
    python sonar_propagate_status.py \\
        --url http://sonarqube.example.com \\
        --username admin --password secret \\
        --project MY_PROJECT

    # Dry run against a list of projects
    python sonar_propagate_status.py \\
        --url http://sonarqube.example.com \\
        --username admin --password secret \\
        --projects projects.txt \\
        --dry-run

Requirements:
    pip install httpx openpyxl tqdm python-dotenv

Notes:
  - Requires a user account with Administer Issues permission on each project.
  - SonarQube caps issue pagination at 10,000 results per query.
    Projects/branches exceeding this will log a warning.
  - Source branch defaults to the isMain branch (isMain: true in the API).
    Override with --source-branch.
"""

import asyncio
import httpx
import argparse
import sys
import os
from collections import defaultdict
from dataclasses import dataclass
from typing import Optional, Any
from datetime import datetime, timezone

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from tqdm import tqdm
from tqdm.asyncio import tqdm as atqdm

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ──────────────────────────────────────────────────────────────────

PAGE_SIZE           = 500
DEFAULT_CONCURRENCY = 10
STALE_DAYS          = 60

# Source statuses to read + the transition to apply on the target
TRANSITION_MAP: dict[str, str] = {
    "CONFIRMED":      "confirm",
    "FALSE_POSITIVE": "falsepositive",
    "ACCEPTED":       "accept",
}

SOURCE_STATUSES = ",".join(TRANSITION_MAP.keys())


# ── Data Models ────────────────────────────────────────────────────────────────

@dataclass
class PropagateResult:
    project_key:    str
    project_name:   str
    source_branch:  str
    target_branch:  str
    issue_key:      str
    component:      str
    rule:           str
    message:        str
    source_status:  str
    transition:     str
    outcome:        str          # "synced" | "failed" | "dry_run" | "skipped"
    error:          Optional[str] = None


# ── SonarQube Client ───────────────────────────────────────────────────────────

class SonarQubeClient:
    def __init__(self, base_url: str, username: str, password: str,
                 concurrency: int = DEFAULT_CONCURRENCY,
                 verify_ssl: bool = True):
        self.base_url   = base_url.rstrip("/")
        self.auth       = (username, password)
        self.verify_ssl = verify_ssl
        self._sem       = asyncio.Semaphore(concurrency)

    async def _get(self, client: httpx.AsyncClient, path: str,
                   params: dict[str, Any] | None = None) -> dict:
        async with self._sem:
            resp = await client.get(
                f"{self.base_url}{path}",
                params=params or {},
                auth=self.auth,
                timeout=60.0,
            )
            resp.raise_for_status()
            return resp.json()

    async def _post(self, client: httpx.AsyncClient, path: str,
                    data: dict[str, Any] | None = None) -> dict:
        async with self._sem:
            resp = await client.post(
                f"{self.base_url}{path}",
                data=data or {},
                auth=self.auth,
                timeout=60.0,
            )
            resp.raise_for_status()
            return resp.json()

    async def get_project_name(self, client: httpx.AsyncClient,
                                project_key: str) -> str:
        try:
            data = await self._get(client, "/api/projects/search",
                                   {"projects": project_key, "ps": 1})
            comps = data.get("components", [])
            return comps[0]["name"] if comps else project_key
        except Exception:
            return project_key

    async def get_branches(self, client: httpx.AsyncClient,
                           project_key: str) -> list[dict]:
        data = await self._get(client, "/api/project_branches/list",
                               {"project": project_key})
        return data.get("branches", [])

    async def get_issues_page(self, client: httpx.AsyncClient,
                               project_key: str, branch: str,
                               statuses: str, page: int) -> dict:
        return await self._get(client, "/api/issues/search", {
            "componentKeys":           project_key,
            "impactSoftwareQualities": "SECURITY",
            "issueStatuses":           statuses,
            "branch":                  branch,
            "ps":                      PAGE_SIZE,
            "p":                       page,
            "additionalFields":        "none",
        })

    async def get_all_issues(self, client: httpx.AsyncClient,
                              project_key: str, branch: str,
                              statuses: str) -> list[dict]:
        """Paginate all issues for a branch+status filter (max 10k)."""
        issues: list[dict] = []
        page = 1
        while True:
            data  = await self.get_issues_page(client, project_key,
                                               branch, statuses, page)
            batch = data.get("issues", [])
            issues.extend(batch)
            pg    = data.get("paging", {})
            total = pg.get("total", 0)
            if page == 1 and total > 10_000:
                _warn(f"[{project_key}/{branch}] {total} issues found; "
                      f"SonarQube caps pagination at 10,000 — "
                      f"only first 10,000 processed.")
            if pg.get("pageIndex", 1) * pg.get("pageSize", PAGE_SIZE) >= min(total, 10_000):
                break
            page += 1
        return issues

    async def do_transition(self, client: httpx.AsyncClient,
                             issue_key: str, transition: str) -> None:
        await self._post(client, "/api/issues/do_transition", {
            "issue":      issue_key,
            "transition": transition,
        })


# ── Helpers ────────────────────────────────────────────────────────────────────

def _warn(msg: str) -> None:
    tqdm.write(f"  [WARN] {msg}", file=sys.stderr)


def _is_stale(analysis_date: Optional[str]) -> bool:
    if not analysis_date:
        return True
    try:
        dt  = datetime.fromisoformat(analysis_date.replace("Z", "+00:00"))
        age = (datetime.now(timezone.utc) - dt).days
        return age > STALE_DAYS
    except Exception:
        return True


def _match_key(issue: dict) -> tuple[str, str, str]:
    """Return (rule, component, hash) — the cross-branch identity of an issue."""
    return (
        issue.get("rule", ""),
        issue.get("component", ""),
        issue.get("hash", ""),
    )


def _issue_status(issue: dict) -> str:
    return issue.get("issueStatus") or issue.get("status", "")


def load_project_keys(path: str) -> list[str]:
    with open(path, encoding="utf-8") as f:
        return [
            line.strip()
            for line in f
            if line.strip() and not line.strip().startswith("#")
        ]


def resolve_project_keys(args: argparse.Namespace) -> list[str]:
    keys: list[str] = []
    if args.project:
        keys.extend(k.strip() for k in args.project.split(",") if k.strip())
    if args.projects:
        keys.extend(load_project_keys(args.projects))
    # deduplicate preserving order
    seen: set[str] = set()
    result: list[str] = []
    for k in keys:
        if k not in seen:
            seen.add(k)
            result.append(k)
    return result


# ── Core logic ─────────────────────────────────────────────────────────────────

async def process_project(
    sonar: SonarQubeClient,
    client: httpx.AsyncClient,
    project_key: str,
    project_name: str,
    branches: list[dict],
    source_branch_override: Optional[str],
    dry_run: bool,
    pbar_scan: tqdm,
    pbar_apply: tqdm,
) -> list[PropagateResult]:

    # ── Identify source branch ────────────────────────────────────────────────
    if source_branch_override:
        source = next(
            (b for b in branches if b["name"] == source_branch_override), None
        )
        if not source:
            _warn(f"[{project_key}] source branch '{source_branch_override}' "
                  f"not found — skipping project.")
            return []
    else:
        source = next((b for b in branches if b.get("isMain")), None)
        if not source:
            _warn(f"[{project_key}] no isMain branch found — skipping project.")
            return []

    source_name = source["name"]

    if not source.get("analysisDate"):
        _warn(f"[{project_key}] source branch '{source_name}' has never been "
              f"scanned — skipping project.")
        return []

    # ── Identify valid target branches ────────────────────────────────────────
    targets = [
        b for b in branches
        if b["name"] != source_name and not _is_stale(b.get("analysisDate"))
    ]
    stale_count = sum(
        1 for b in branches
        if b["name"] != source_name and _is_stale(b.get("analysisDate"))
    )
    if stale_count:
        _warn(f"[{project_key}] {stale_count} branch(es) skipped — "
              f"no scan in last {STALE_DAYS} days.")

    if not targets:
        tqdm.write(f"  [{project_key}] no active target branches found.")
        return []

    # ── Phase: fetch source acknowledged issues ───────────────────────────────
    try:
        source_issues = await sonar.get_all_issues(
            client, project_key, source_name, SOURCE_STATUSES
        )
    except Exception as exc:
        _warn(f"[{project_key}] failed to fetch source issues: {exc}")
        return []
    finally:
        pbar_scan.update(1)

    # Build match index: {(rule, component, hash) → (issue_key, status)}
    # If duplicate keys exist, last one wins (extremely rare)
    source_index: dict[tuple, tuple[str, str]] = {}
    for issue in source_issues:
        mk = _match_key(issue)
        if mk[2]:  # only index issues that have a hash
            source_index[mk] = (issue["key"], _issue_status(issue))

    if not source_index:
        tqdm.write(f"  [{project_key}] source branch '{source_name}' has no "
                   f"acknowledged issues to propagate.")
        return []

    # ── Phase: fetch OPEN issues on each target branch + match ───────────────
    results: list[PropagateResult] = []

    async def process_target(target_branch: str) -> list[PropagateResult]:
        branch_results: list[PropagateResult] = []
        try:
            target_issues = await sonar.get_all_issues(
                client, project_key, target_branch, "OPEN"
            )
        except Exception as exc:
            _warn(f"[{project_key}/{target_branch}] failed to fetch issues: {exc}")
            return []
        finally:
            pbar_scan.update(1)

        for issue in target_issues:
            mk = _match_key(issue)
            if not mk[2]:
                continue  # no hash — can't match reliably
            if mk not in source_index:
                continue  # no matching acknowledged issue on source

            _, source_status = source_index[mk]
            transition = TRANSITION_MAP.get(source_status)
            if not transition:
                continue

            branch_results.append(PropagateResult(
                project_key=project_key,
                project_name=project_name,
                source_branch=source_name,
                target_branch=target_branch,
                issue_key=issue["key"],
                component=issue.get("component", ""),
                rule=issue.get("rule", ""),
                message=issue.get("message", ""),
                source_status=source_status,
                transition=transition,
                outcome="pending",
            ))

        return branch_results

    target_batches = await asyncio.gather(
        *[process_target(t["name"]) for t in targets],
        return_exceptions=True,
    )
    for batch in target_batches:
        if isinstance(batch, Exception):
            _warn(f"[{project_key}] unexpected error processing a target branch: {batch}")
        else:
            results.extend(batch)

    # ── Phase: apply transitions ──────────────────────────────────────────────
    async def apply(r: PropagateResult) -> PropagateResult:
        if dry_run:
            pbar_apply.update(1)
            r.outcome = "dry_run"
            return r
        try:
            await sonar.do_transition(client, r.issue_key, r.transition)
            r.outcome = "synced"
        except Exception as exc:
            r.outcome = "failed"
            r.error   = str(exc)
        finally:
            pbar_apply.update(1)
        return r

    results = list(await asyncio.gather(*[apply(r) for r in results]))
    return results


# ── Orchestration ──────────────────────────────────────────────────────────────

async def run(args: argparse.Namespace) -> list[PropagateResult]:
    project_keys = resolve_project_keys(args)
    if not project_keys:
        print("Error: no project keys provided. "
              "Use --project and/or --projects.", file=sys.stderr)
        sys.exit(1)

    sonar = SonarQubeClient(
        base_url=args.url,
        username=args.username,
        password=args.password,
        concurrency=args.concurrency,
        verify_ssl=not args.no_verify_ssl,
    )

    async with httpx.AsyncClient(verify=not args.no_verify_ssl) as client:

        # ── Phase 1: resolve project names + branches ──────────────────────
        async def fetch_project_info(key: str) -> dict:
            name, branches = await asyncio.gather(
                sonar.get_project_name(client, key),
                sonar.get_branches(client, key),
            )
            return {"key": key, "name": name, "branches": branches}

        projects: list[dict] = await atqdm.gather(
            *[fetch_project_info(k) for k in project_keys],
            desc="Fetching project info",
            unit="project",
            total=len(project_keys),
        )

        total_branches = sum(len(p["branches"]) for p in projects)
        tqdm.write(f"\n{len(projects)} project(s), "
                   f"{total_branches} branch(es) total.\n")

        # ── Phase 2 + 3: scan branches and apply transitions ───────────────
        # scan ticks: 1 for source + N for targets per project
        total_scan_ticks = sum(
            1 + sum(
                1 for b in p["branches"]
                if b["name"] != (
                    next((b2["name"] for b2 in p["branches"]
                          if b2.get("isMain")), None)
                    if not args.source_branch
                    else args.source_branch
                )
                and not _is_stale(b.get("analysisDate"))
            )
            for p in projects
        )

        all_results: list[PropagateResult] = []

        with tqdm(desc="Scanning branches",  unit="branch",
                  total=total_scan_ticks) as pbar_scan, \
             tqdm(desc="Applying transitions", unit="issue",
                  total=0) as pbar_apply:

            for proj in projects:
                proj_results = await process_project(
                    sonar, client,
                    proj["key"], proj["name"], proj["branches"],
                    args.source_branch or None,
                    args.dry_run,
                    pbar_scan,
                    pbar_apply,
                )
                # update total on pbar_apply now we know how many to apply
                pbar_apply.total = (pbar_apply.total or 0) + len(proj_results)
                pbar_apply.refresh()
                all_results.extend(proj_results)

    return all_results


# ── Console summary ────────────────────────────────────────────────────────────

def print_summary(results: list[PropagateResult]) -> None:
    from collections import Counter
    counts = Counter(r.outcome for r in results)
    print("\n" + "═" * 60)
    print("  PROPAGATION SUMMARY")
    print("═" * 60)
    print(f"  Total matched    : {len(results)}")
    print(f"  Synced           : {counts.get('synced', 0)}")
    print(f"  Failed           : {counts.get('failed', 0)}")
    print(f"  Dry run          : {counts.get('dry_run', 0)}")
    print("═" * 60)


# ── Excel export ───────────────────────────────────────────────────────────────

_HDR_FILL    = PatternFill("solid", fgColor="1F3864")
_SYNC_FILL   = PatternFill("solid", fgColor="E2EFDA")   # green  — synced
_FAIL_FILL   = PatternFill("solid", fgColor="FCE4D6")   # orange — failed
_DRY_FILL    = PatternFill("solid", fgColor="DEEAF1")   # blue   — dry run
_SKIP_FILL   = PatternFill("solid", fgColor="F2F2F2")   # grey   — skipped
_ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_TOTAL_FILL  = PatternFill("solid", fgColor="375623")

_HDR_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT   = Font(name="Calibri", size=10)
_BOLD_FONT   = Font(bold=True, name="Calibri", size=10)
_TOTAL_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

_THIN        = Side(border_style="thin", color="B8CCE4")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center")
_RIGHT       = Alignment(horizontal="right",  vertical="center")


def _hdr(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _HDR_FILL; c.font = _HDR_FONT
    c.alignment = _CENTER; c.border = _BORDER


def _cell(ws, row: int, col: int, value,
          fill=None, align=None, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill or _WHITE_FILL
    c.font      = _BOLD_FONT if bold else _BODY_FONT
    c.alignment = align or _LEFT
    c.border    = _BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _outcome_fill(outcome: str) -> PatternFill:
    return {
        "synced":  _SYNC_FILL,
        "failed":  _FAIL_FILL,
        "dry_run": _DRY_FILL,
        "skipped": _SKIP_FILL,
    }.get(outcome, _WHITE_FILL)


def _build_summary_sheet(ws, results: list[PropagateResult]) -> None:
    ws.title = "Summary"
    ws.freeze_panes = "A2"

    headers = ["Project Key", "Project Name", "Source Branch", "Target Branch",
               "Synced", "Failed", "Dry Run", "Total"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    # aggregate per (project_key, source_branch, target_branch)
    stats: dict[tuple, dict] = defaultdict(
        lambda: {"name": "", "s": 0, "f": 0, "d": 0}
    )
    for r in results:
        k = (r.project_key, r.source_branch, r.target_branch)
        stats[k]["name"] = r.project_name
        if r.outcome == "synced":   stats[k]["s"] += 1
        elif r.outcome == "failed": stats[k]["f"] += 1
        elif r.outcome == "dry_run":stats[k]["d"] += 1

    for row_idx, ((proj_key, src, tgt), s) in enumerate(sorted(stats.items()), 2):
        fill  = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        total = s["s"] + s["f"] + s["d"]
        _cell(ws, row_idx, 1, proj_key,  fill=fill)
        _cell(ws, row_idx, 2, s["name"], fill=fill)
        _cell(ws, row_idx, 3, src,       fill=fill)
        _cell(ws, row_idx, 4, tgt,       fill=fill)
        _cell(ws, row_idx, 5, s["s"],    fill=fill, align=_RIGHT)
        _cell(ws, row_idx, 6, s["f"],    fill=fill, align=_RIGHT)
        _cell(ws, row_idx, 7, s["d"],    fill=fill, align=_RIGHT)
        _cell(ws, row_idx, 8, total,     fill=fill, align=_RIGHT, bold=True)

    total_row = len(stats) + 2
    t = sum(s["s"] for s in stats.values())
    f = sum(s["f"] for s in stats.values())
    d = sum(s["d"] for s in stats.values())
    _cell(ws, total_row, 1, f"TOTAL  ({len(stats)} branch(es))",
          fill=_TOTAL_FILL, bold=True)
    ws.cell(row=total_row, column=1).font = _TOTAL_FONT
    for col in [2, 3, 4]:
        ws.cell(row=total_row, column=col).fill   = _TOTAL_FILL
        ws.cell(row=total_row, column=col).border = _BORDER
    for col, val in [(5, t), (6, f), (7, d), (8, t + f + d)]:
        _cell(ws, total_row, col, val, fill=_TOTAL_FILL, align=_RIGHT, bold=True)
        ws.cell(row=total_row, column=col).font = _TOTAL_FONT

    _auto_width(ws)


def _build_detail_sheet(ws, results: list[PropagateResult]) -> None:
    ws.title = "Detail"
    ws.freeze_panes = "A2"

    headers = ["Project Key", "Project Name", "Source Branch", "Target Branch",
               "Issue Key", "Component", "Rule", "Message",
               "Source Status", "Transition Applied", "Outcome", "Error"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    for row_idx, r in enumerate(results, 2):
        fill = _outcome_fill(r.outcome)
        _cell(ws, row_idx,  1, r.project_key,   fill=fill)
        _cell(ws, row_idx,  2, r.project_name,  fill=fill)
        _cell(ws, row_idx,  3, r.source_branch, fill=fill)
        _cell(ws, row_idx,  4, r.target_branch, fill=fill)
        _cell(ws, row_idx,  5, r.issue_key,     fill=fill)
        _cell(ws, row_idx,  6, r.component,     fill=fill)
        _cell(ws, row_idx,  7, r.rule,          fill=fill)
        _cell(ws, row_idx,  8, r.message,       fill=fill)
        _cell(ws, row_idx,  9, r.source_status, fill=fill, align=_CENTER)
        _cell(ws, row_idx, 10, r.transition,    fill=fill, align=_CENTER)
        _cell(ws, row_idx, 11, r.outcome,       fill=fill, align=_CENTER, bold=True)
        _cell(ws, row_idx, 12, r.error or "",   fill=fill)

    _auto_width(ws)


def export_excel(results: list[PropagateResult], path: str) -> None:
    wb = Workbook()
    _build_summary_sheet(wb.active, results)
    _build_detail_sheet(wb.create_sheet(), results)
    wb.save(path)


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Propagate acknowledged issue statuses from a source branch "
                    "to matching OPEN issues on all other active branches.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--url",      default=os.getenv("SONAR_URL"),
                   help="SonarQube base URL  [env: SONAR_URL]")
    p.add_argument("--username", default=os.getenv("SONAR_USERNAME"),
                   help="SonarQube username  [env: SONAR_USERNAME]")
    p.add_argument("--password", default=os.getenv("SONAR_PASSWORD"),
                   help="SonarQube password  [env: SONAR_PASSWORD]")
    p.add_argument("--project",  default=os.getenv("SONAR_PROJECT"),
                   help="Single project key (comma-separated for multiple)  "
                        "[env: SONAR_PROJECT]")
    p.add_argument("--projects", default=os.getenv("SONAR_PROJECTS_FILE"),
                   help="Text file with project keys, one per line  "
                        "[env: SONAR_PROJECTS_FILE]")
    p.add_argument("--source-branch", default=os.getenv("SONAR_SOURCE_BRANCH"),
                   help="Branch to read statuses from (default: isMain branch)  "
                        "[env: SONAR_SOURCE_BRANCH]")
    p.add_argument("--output",   default=os.getenv("SONAR_PROPAGATE_OUTPUT",
                                                    "propagate_report.xlsx"),
                   help="Excel report output path  [env: SONAR_PROPAGATE_OUTPUT]")
    p.add_argument("--concurrency", type=int,
                   default=int(os.getenv("SONAR_CONCURRENCY", DEFAULT_CONCURRENCY)),
                   help="Max parallel API requests  [env: SONAR_CONCURRENCY]")
    p.add_argument("--dry-run", action="store_true",
                   help="Scan and report without applying any transitions")
    p.add_argument("--no-verify-ssl", action="store_true",
                   help="Disable SSL certificate verification (self-signed certs)")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    missing = [f"--{k}" for k, v in
               [("url", args.url), ("username", args.username),
                ("password", args.password)] if not v]
    if missing:
        print(f"Error: missing required values: {', '.join(missing)}\n"
              f"Pass them as CLI args or set SONAR_URL / SONAR_USERNAME / "
              f"SONAR_PASSWORD in your environment / .env file.",
              file=sys.stderr)
        sys.exit(1)

    if not args.project and not args.projects:
        print("Error: provide at least one of --project or --projects.",
              file=sys.stderr)
        sys.exit(1)

    print("SonarQube Status Propagator")
    print(f"  URL           : {args.url}")
    print(f"  Username      : {args.username}")
    print(f"  Project       : {args.project or '—'}")
    print(f"  Projects file : {args.projects or '—'}")
    print(f"  Source branch : {args.source_branch or 'isMain (auto)'}")
    print(f"  Concurrency   : {args.concurrency}")
    print(f"  Output        : {args.output}")
    print(f"  Dry run       : {args.dry_run}")
    print(f"  Verify SSL    : {not args.no_verify_ssl}")
    print()

    results = asyncio.run(run(args))

    if not results:
        print("No matching issues found to propagate.")
        return

    print_summary(results)
    export_excel(results, args.output)
    print(f"\nExcel report saved → {args.output}")


if __name__ == "__main__":
    main()
