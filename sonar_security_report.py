#!/usr/bin/env python3
"""
SonarQube Security Report
=========================
Fetches vulnerability and security hotspot counts across all projects,
branches, and pull requests in SonarQube Enterprise (2025.1 LTA).

Outputs:
  - Summary table  : project name, branch count, PR count, last scan, totals
  - Detailed report: per project → per branch/PR → vuln & hotspot counts by status/severity
  - JSON file      : full structured data
  - Excel file     : 3 formatted sheets — Summary, Branches, Pull Requests

Usage:
    python sonar_security_report.py \
        --url   http://sonarqube.example.com \
        --username admin \
        --password secret

    # Custom output paths, limit concurrency
    python sonar_security_report.py \
        --url http://sonar.internal \
        --username admin --password secret \
        --output /tmp/report.json \
        --excel  /tmp/report.xlsx \
        --concurrency 5 \
        --summary-only

    # Skip SSL verification (self-signed certs)
    python sonar_security_report.py ... --no-verify-ssl

Requirements:
    pip install httpx openpyxl
"""

import asyncio
import httpx
import json
import argparse
import sys
import os
from dataclasses import dataclass, field
from typing import Optional, Any
from datetime import datetime, timezone

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed; fall back to system environment

from tqdm import tqdm
from tqdm.asyncio import tqdm as atqdm

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────────

PAGE_SIZE = 500
DEFAULT_CONCURRENCY = 10

# Hotspot review statuses in SonarQube
HOTSPOT_STATUSES = ["TO_REVIEW", "ACKNOWLEDGED", "FIXED", "SAFE"]


# ── Data Models ───────────────────────────────────────────────────────────────

@dataclass
class VulnerabilityStats:
    total: int = 0
    by_status: dict[str, int] = field(default_factory=dict)
    by_severity: dict[str, int] = field(default_factory=dict)


@dataclass
class HotspotStats:
    total: int = 0
    by_status: dict[str, int] = field(default_factory=dict)


@dataclass
class BranchReport:
    name: str
    branch_type: str          # MAIN, LONG, SHORT
    is_main: bool
    last_analysis: Optional[str]
    vulnerabilities: VulnerabilityStats = field(default_factory=VulnerabilityStats)
    hotspots: HotspotStats = field(default_factory=HotspotStats)


@dataclass
class PullRequestReport:
    key: str
    title: str
    source_branch: str
    target_branch: str
    last_analysis: Optional[str]
    vulnerabilities: VulnerabilityStats = field(default_factory=VulnerabilityStats)
    hotspots: HotspotStats = field(default_factory=HotspotStats)


@dataclass
class ProjectReport:
    key: str
    name: str
    last_analysis: Optional[str]
    branches: list[BranchReport] = field(default_factory=list)
    pull_requests: list[PullRequestReport] = field(default_factory=list)

    def total_vulns(self) -> int:
        return (sum(b.vulnerabilities.total for b in self.branches)
                + sum(p.vulnerabilities.total for p in self.pull_requests))

    def total_hotspots(self) -> int:
        return (sum(b.hotspots.total for b in self.branches)
                + sum(p.hotspots.total for p in self.pull_requests))


# ── SonarQube Async Client ────────────────────────────────────────────────────

class SonarQubeClient:
    def __init__(self, base_url: str, username: str, password: str,
                 concurrency: int = DEFAULT_CONCURRENCY, verify_ssl: bool = True):
        self.base_url = base_url.rstrip("/")
        self.auth = (username, password)
        self.verify_ssl = verify_ssl
        self._sem = asyncio.Semaphore(concurrency)

    async def _get(self, client: httpx.AsyncClient, path: str,
                   params: dict[str, Any] | None = None) -> dict:
        async with self._sem:
            resp = await client.get(
                f"{self.base_url}{path}",
                params=params or {},
                auth=self.auth,
                timeout=60.0,
            )
            resp.raise_for_status()
            return resp.json()

    # ── Projects ──────────────────────────────────────────────────────────────

    async def get_all_projects(self, client: httpx.AsyncClient) -> list[dict]:
        """Fetch every project using pagination."""
        projects, page = [], 1
        pbar = tqdm(desc="Fetching projects", unit="project")
        try:
            while True:
                data = await self._get(client, "/api/projects/search",
                                       {"ps": PAGE_SIZE, "p": page})
                batch = data.get("components", [])
                projects.extend(batch)
                pg = data.get("paging", {})
                if page == 1:
                    pbar.total = pg.get("total", 0)
                    pbar.refresh()
                pbar.update(len(batch))
                if pg.get("pageIndex", 1) * pg.get("pageSize", PAGE_SIZE) >= pg.get("total", 0):
                    break
                page += 1
        finally:
            pbar.close()
        return projects

    # ── Branches ──────────────────────────────────────────────────────────────

    async def get_branches(self, client: httpx.AsyncClient, project_key: str) -> list[dict]:
        data = await self._get(client, "/api/project_branches/list",
                               {"project": project_key})
        return data.get("branches", [])

    # ── Pull Requests ─────────────────────────────────────────────────────────

    async def get_pull_requests(self, client: httpx.AsyncClient,
                                project_key: str) -> list[dict]:
        """Returns PRs for the project; empty list if the project has none or
        PR analysis is not configured (404 / 400 response)."""
        try:
            data = await self._get(client, "/api/project_pull_requests/list",
                                   {"project": project_key})
            return data.get("pullRequests", [])
        except httpx.HTTPStatusError as exc:
            if exc.response.status_code in (400, 404):
                return []
            raise

    # ── Vulnerabilities ───────────────────────────────────────────────────────

    async def get_vulnerability_stats(
        self, client: httpx.AsyncClient, project_key: str,
        branch: str | None = None, pull_request: str | None = None,
    ) -> VulnerabilityStats:
        """One API call; uses facets to get status + severity breakdown."""
        params: dict[str, Any] = {
            "types": "VULNERABILITY",
            "componentKeys": project_key,
            "ps": 1,
            "facets": "statuses,severities",
        }
        if branch:
            params["branch"] = branch
        if pull_request:
            params["pullRequest"] = pull_request

        data = await self._get(client, "/api/issues/search", params)

        total = data.get("paging", {}).get("total", 0)
        by_status: dict[str, int] = {}
        by_severity: dict[str, int] = {}

        for facet in data.get("facets", []):
            if facet["property"] == "statuses":
                by_status = {v["val"]: v["count"]
                             for v in facet["values"] if v["count"] > 0}
            elif facet["property"] == "severities":
                by_severity = {v["val"]: v["count"]
                               for v in facet["values"] if v["count"] > 0}

        return VulnerabilityStats(total=total, by_status=by_status,
                                  by_severity=by_severity)

    # ── Hotspots ──────────────────────────────────────────────────────────────

    async def get_hotspot_stats(
        self, client: httpx.AsyncClient, project_key: str,
        branch: str | None = None, pull_request: str | None = None,
    ) -> HotspotStats:
        """One call per hotspot status (4 total), run concurrently."""
        async def fetch_status(status: str) -> tuple[str, int]:
            params: dict[str, Any] = {
                "projectKey": project_key,
                "status": status,
                "ps": 1,
            }
            if branch:
                params["branch"] = branch
            if pull_request:
                params["pullRequest"] = pull_request
            data = await self._get(client, "/api/hotspots/search", params)
            return status, data.get("paging", {}).get("total", 0)

        results = await asyncio.gather(
            *[fetch_status(s) for s in HOTSPOT_STATUSES],
            return_exceptions=True,
        )

        by_status: dict[str, int] = {}
        total = 0
        for i, res in enumerate(results):
            status = HOTSPOT_STATUSES[i]
            if isinstance(res, Exception):
                _warn(f"hotspot fetch failed [project={project_key} "
                      f"branch={branch or pull_request} status={status}]: {res}")
                by_status[status] = 0
            else:
                _, count = res
                by_status[status] = count
                total += count

        return HotspotStats(total=total, by_status=by_status)


# ── Orchestration ─────────────────────────────────────────────────────────────

def _warn(msg: str) -> None:
    print(f"  [WARN] {msg}", file=sys.stderr, flush=True)


async def process_branch(sonar: SonarQubeClient, client: httpx.AsyncClient,
                         project_key: str, branch: dict) -> BranchReport:
    name = branch["name"]
    vuln_stats, hotspot_stats = await asyncio.gather(
        sonar.get_vulnerability_stats(client, project_key, branch=name),
        sonar.get_hotspot_stats(client, project_key, branch=name),
    )
    return BranchReport(
        name=name,
        branch_type=branch.get("type", "UNKNOWN"),
        is_main=branch.get("isMain", False),
        last_analysis=branch.get("analysisDate"),
        vulnerabilities=vuln_stats,
        hotspots=hotspot_stats,
    )


async def process_pull_request(sonar: SonarQubeClient, client: httpx.AsyncClient,
                               project_key: str, pr: dict) -> PullRequestReport:
    key = str(pr["key"])
    vuln_stats, hotspot_stats = await asyncio.gather(
        sonar.get_vulnerability_stats(client, project_key, pull_request=key),
        sonar.get_hotspot_stats(client, project_key, pull_request=key),
    )
    return PullRequestReport(
        key=key,
        title=pr.get("title", ""),
        source_branch=pr.get("branch", ""),
        target_branch=pr.get("base", ""),
        last_analysis=pr.get("analysisDate"),
        vulnerabilities=vuln_stats,
        hotspots=hotspot_stats,
    )


async def process_project(sonar: SonarQubeClient, client: httpx.AsyncClient,
                          project: dict, branches_raw: list[dict],
                          prs_raw: list[dict],
                          pbar: Optional[tqdm] = None) -> ProjectReport:
    key = project["key"]

    async def do_branch(b: dict):
        try:
            return await process_branch(sonar, client, key, b)
        except Exception as exc:
            return exc
        finally:
            if pbar:
                pbar.update(1)

    async def do_pr(pr: dict):
        try:
            return await process_pull_request(sonar, client, key, pr)
        except Exception as exc:
            return exc
        finally:
            if pbar:
                pbar.update(1)

    nb = len(branches_raw)
    all_results = await asyncio.gather(
        *[do_branch(b) for b in branches_raw],
        *[do_pr(pr) for pr in prs_raw],
    )

    branch_reports: list[BranchReport] = []
    pr_reports: list[PullRequestReport] = []

    for i, result in enumerate(all_results):
        if isinstance(result, Exception):
            label = branches_raw[i]["name"] if i < nb else prs_raw[i - nb]["key"]
            _warn(f"[{key}] failed on {'branch' if i < nb else 'PR'} "
                  f"'{label}': {result}")
        elif i < nb:
            branch_reports.append(result)
        else:
            pr_reports.append(result)

    return ProjectReport(
        key=key,
        name=project["name"],
        last_analysis=project.get("lastAnalysisDate"),
        branches=branch_reports,
        pull_requests=pr_reports,
    )


async def run(args: argparse.Namespace) -> list[ProjectReport]:
    sonar = SonarQubeClient(
        base_url=args.url,
        username=args.username,
        password=args.password,
        concurrency=args.concurrency,
        verify_ssl=not args.no_verify_ssl,
    )
    async with httpx.AsyncClient(verify=not args.no_verify_ssl) as client:

        # ── Phase 1: project list (progress bar inside get_all_projects) ──────
        projects = await sonar.get_all_projects(client)
        print(f"\nFound {len(projects)} project(s).\n", flush=True)

        # ── Phase 2: branch & PR lists for every project ──────────────────────
        async def fetch_lists(p: dict) -> tuple[list[dict], list[dict]]:
            branches, prs = await asyncio.gather(
                sonar.get_branches(client, p["key"]),
                sonar.get_pull_requests(client, p["key"]),
            )
            return branches, prs

        list_results: list[tuple[list[dict], list[dict]]] = await atqdm.gather(
            *[fetch_lists(p) for p in projects],
            desc="Fetching branches & PRs",
            unit="project",
            total=len(projects),
        )

        # ── Phase 3: security data for every branch / PR ──────────────────────
        total_items = sum(len(b) + len(p) for b, p in list_results)
        with tqdm(desc="Fetching security data", unit="branch/PR",
                  total=total_items) as pbar:
            results = await asyncio.gather(
                *[
                    process_project(sonar, client, proj, branches, prs, pbar)
                    for proj, (branches, prs) in zip(projects, list_results)
                ],
                return_exceptions=True,
            )

    reports: list[ProjectReport] = []
    for i, r in enumerate(results):
        if isinstance(r, Exception):
            _warn(f"project '{projects[i]['key']}' failed entirely: {r}")
        else:
            reports.append(r)

    return reports


# ── Output helpers ────────────────────────────────────────────────────────────

def _fmt_date(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return iso


def _table(headers: list[str], rows: list[list[str]]) -> str:
    widths = [
        max(len(h), max((len(r[i]) for r in rows), default=0))
        for i, h in enumerate(headers)
    ]
    sep = "+-" + "-+-".join("-" * w for w in widths) + "-+"
    def fmt_row(cells):
        return "| " + " | ".join(c.ljust(widths[i]) for i, c in enumerate(cells)) + " |"
    lines = [sep, fmt_row(headers), sep.replace("-", "=")]
    for row in rows:
        lines.append(fmt_row(row))
    lines.append(sep)
    return "\n".join(lines)


def print_summary(reports: list[ProjectReport]) -> None:
    rows = []
    for r in sorted(reports, key=lambda x: x.name.lower()):
        rows.append([
            r.name,
            r.key,
            str(len(r.branches)),
            str(len(r.pull_requests)),
            str(r.total_vulns()),
            str(r.total_hotspots()),
            _fmt_date(r.last_analysis),
        ])
    headers = ["Project Name", "Key", "Branches", "PRs",
                "Vulnerabilities", "Hotspots", "Last Scan"]
    print("\n" + "═" * 80)
    print("  PROJECT SUMMARY")
    print("═" * 80)
    print(_table(headers, rows))


def print_detail(reports: list[ProjectReport]) -> None:
    print("\n" + "═" * 80)
    print("  DETAILED SECURITY REPORT")
    print("═" * 80)

    for project in sorted(reports, key=lambda x: x.name.lower()):
        print(f"\n┌─ PROJECT : {project.name}  [{project.key}]")
        print(f"│  Last scan         : {_fmt_date(project.last_analysis)}")
        print(f"│  Branches          : {len(project.branches)}")
        print(f"│  Pull Requests     : {len(project.pull_requests)}")
        print(f"│  Total Vulns       : {project.total_vulns()}")
        print(f"│  Total Hotspots    : {project.total_hotspots()}")
        print("│")

        def _print_item(label: str, last_analysis: Optional[str],
                        vulns: VulnerabilityStats, hotspots: HotspotStats,
                        indent: str = "│    ") -> None:
            print(f"{indent}├─ {label}")
            print(f"{indent}│    Last scan     : {_fmt_date(last_analysis)}")
            # Vulnerabilities
            print(f"{indent}│    Vulnerabilities : {vulns.total}")
            if vulns.by_status:
                kv = "  ".join(f"{k}={v}" for k, v in sorted(vulns.by_status.items()))
                print(f"{indent}│      by status   : {kv}")
            if vulns.by_severity:
                kv = "  ".join(f"{k}={v}" for k, v in sorted(vulns.by_severity.items()))
                print(f"{indent}│      by severity : {kv}")
            # Hotspots
            print(f"{indent}│    Hotspots        : {hotspots.total}")
            if hotspots.by_status:
                kv = "  ".join(f"{k}={v}" for k, v in hotspots.by_status.items())
                print(f"{indent}│      by status   : {kv}")

        if project.branches:
            print("│  ── Branches ──")
            for b in sorted(project.branches,
                            key=lambda x: (not x.is_main, x.name.lower())):
                tag = "[MAIN]" if b.is_main else f"[{b.branch_type}]"
                _print_item(f"BRANCH {tag}  {b.name}", b.last_analysis,
                            b.vulnerabilities, b.hotspots)

        if project.pull_requests:
            print("│  ── Pull Requests ──")
            for pr in sorted(project.pull_requests, key=lambda x: x.key):
                label = (f"PR #{pr.key}  \"{pr.title}\""
                         f"  ({pr.source_branch} → {pr.target_branch})")
                _print_item(label, pr.last_analysis,
                            pr.vulnerabilities, pr.hotspots)

        print(f"└{'─' * 78}")


# ── JSON serialisation ────────────────────────────────────────────────────────

def to_dict(reports: list[ProjectReport]) -> dict:
    def branch_dict(b: BranchReport) -> dict:
        return {
            "name": b.name,
            "type": b.branch_type,
            "is_main": b.is_main,
            "last_analysis": b.last_analysis,
            "vulnerabilities": {
                "total": b.vulnerabilities.total,
                "by_status": b.vulnerabilities.by_status,
                "by_severity": b.vulnerabilities.by_severity,
            },
            "hotspots": {
                "total": b.hotspots.total,
                "by_status": b.hotspots.by_status,
            },
        }

    def pr_dict(pr: PullRequestReport) -> dict:
        return {
            "key": pr.key,
            "title": pr.title,
            "source_branch": pr.source_branch,
            "target_branch": pr.target_branch,
            "last_analysis": pr.last_analysis,
            "vulnerabilities": {
                "total": pr.vulnerabilities.total,
                "by_status": pr.vulnerabilities.by_status,
                "by_severity": pr.vulnerabilities.by_severity,
            },
            "hotspots": {
                "total": pr.hotspots.total,
                "by_status": pr.hotspots.by_status,
            },
        }

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_projects": len(reports),
        "grand_total_vulnerabilities": sum(r.total_vulns() for r in reports),
        "grand_total_hotspots": sum(r.total_hotspots() for r in reports),
        "projects": [
            {
                "key": r.key,
                "name": r.name,
                "last_analysis": r.last_analysis,
                "branch_count": len(r.branches),
                "pr_count": len(r.pull_requests),
                "total_vulnerabilities": r.total_vulns(),
                "total_hotspots": r.total_hotspots(),
                "branches": [branch_dict(b) for b in r.branches],
                "pull_requests": [pr_dict(p) for p in r.pull_requests],
            }
            for r in sorted(reports, key=lambda x: x.name.lower())
        ],
    }


# ── Excel export ─────────────────────────────────────────────────────────────

# Preferred column ordering for dynamic status/severity values
_VULN_STATUS_ORDER  = ["OPEN", "CONFIRMED", "REOPENED", "ACCEPTED",
                        "FALSE_POSITIVE", "RESOLVED", "CLOSED"]
_VULN_SEV_ORDER     = ["BLOCKER", "CRITICAL", "MAJOR", "MINOR", "INFO"]

# ── Styles ────────────────────────────────────────────────────────────────────
_HDR_FILL      = PatternFill("solid", fgColor="1F3864")   # dark navy
_GRP_FILL      = PatternFill("solid", fgColor="2E75B6")   # medium blue
_TOTAL_FILL    = PatternFill("solid", fgColor="375623")   # dark green  (grand total row)
_ALT_FILL      = PatternFill("solid", fgColor="DEEAF1")   # light blue  (alternating)
_WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
_WARN_FILL     = PatternFill("solid", fgColor="FCE4D6")   # light orange (>0 hotspots/vulns)

_HDR_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_GRP_FONT      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_TOTAL_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT     = Font(name="Calibri", size=10)
_BOLD_FONT     = Font(bold=True, name="Calibri", size=10)

_THIN          = Side(border_style="thin", color="B8CCE4")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center")
_RIGHT         = Alignment(horizontal="right",  vertical="center")


def _hdr_cell(ws, row: int, col: int, value: str,
               fill=None, font=None, align=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill  or _HDR_FILL
    c.font   = font  or _HDR_FONT
    c.alignment = align or _CENTER
    c.border = _BORDER


def _body_cell(ws, row: int, col: int, value,
               fill=None, align=None, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill  or _WHITE_FILL
    c.font      = _BOLD_FONT if bold else _BODY_FONT
    c.alignment = align or _LEFT
    c.border    = _BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 50) -> None:
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _collect_dynamic_cols(reports: list[ProjectReport]) -> tuple[list[str], list[str]]:
    """Return (sorted vuln statuses, sorted vuln severities) found across all data."""
    statuses: set[str]   = set()
    severities: set[str] = set()

    def _add(stats: VulnerabilityStats) -> None:
        statuses.update(stats.by_status.keys())
        severities.update(stats.by_severity.keys())

    for r in reports:
        for b in r.branches:
            _add(b.vulnerabilities)
        for p in r.pull_requests:
            _add(p.vulnerabilities)

    def _sort(values: set[str], order: list[str]) -> list[str]:
        known   = [v for v in order if v in values]
        unknown = sorted(values - set(order))
        return known + unknown

    return _sort(statuses, _VULN_STATUS_ORDER), _sort(severities, _VULN_SEV_ORDER)


def _parse_iso(iso: Optional[str]) -> Optional[datetime]:
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, reports: list[ProjectReport]) -> None:
    ws.title = "Summary"
    ws.freeze_panes = "A2"

    headers = ["Project Name", "Project Key", "Branches", "Pull Requests",
               "Total Vulnerabilities", "Total Hotspots", "Last Scan"]
    for col, h in enumerate(headers, 1):
        _hdr_cell(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    sorted_reports = sorted(reports, key=lambda x: x.name.lower())
    for row_idx, r in enumerate(sorted_reports, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        _body_cell(ws, row_idx, 1, r.name,              fill=fill, align=_LEFT)
        _body_cell(ws, row_idx, 2, r.key,               fill=fill, align=_LEFT)
        _body_cell(ws, row_idx, 3, len(r.branches),     fill=fill, align=_CENTER)
        _body_cell(ws, row_idx, 4, len(r.pull_requests),fill=fill, align=_CENTER)
        _body_cell(ws, row_idx, 5, r.total_vulns(),     fill=fill, align=_RIGHT)
        _body_cell(ws, row_idx, 6, r.total_hotspots(),  fill=fill, align=_RIGHT)
        _body_cell(ws, row_idx, 7, _parse_iso(r.last_analysis), fill=fill, align=_CENTER)
        ws.cell(row=row_idx, column=7).number_format = "YYYY-MM-DD HH:MM"

    # Grand total row
    total_row = len(sorted_reports) + 2
    _body_cell(ws, total_row, 1, f"TOTAL  ({len(reports)} projects)",
               fill=_TOTAL_FILL, align=_LEFT, bold=True)
    ws.cell(row=total_row, column=1).font = _TOTAL_FONT
    for col in range(2, 8):
        ws.cell(row=total_row, column=col).fill = _TOTAL_FILL
        ws.cell(row=total_row, column=col).border = _BORDER
    _body_cell(ws, total_row, 3, sum(len(r.branches)     for r in reports),
               fill=_TOTAL_FILL, align=_CENTER, bold=True)
    ws.cell(row=total_row, column=3).font = _TOTAL_FONT
    _body_cell(ws, total_row, 4, sum(len(r.pull_requests) for r in reports),
               fill=_TOTAL_FILL, align=_CENTER, bold=True)
    ws.cell(row=total_row, column=4).font = _TOTAL_FONT
    _body_cell(ws, total_row, 5, sum(r.total_vulns()    for r in reports),
               fill=_TOTAL_FILL, align=_RIGHT, bold=True)
    ws.cell(row=total_row, column=5).font = _TOTAL_FONT
    _body_cell(ws, total_row, 6, sum(r.total_hotspots() for r in reports),
               fill=_TOTAL_FILL, align=_RIGHT, bold=True)
    ws.cell(row=total_row, column=6).font = _TOTAL_FONT

    _auto_width(ws)


def _write_security_sheet(
    ws,
    title: str,
    fixed_headers: list[str],
    rows: list[dict],
    vuln_statuses: list[str],
) -> None:
    """Generic writer used for both Branches and Pull Requests sheets."""
    ws.title = title
    ws.freeze_panes = "A3"

    # ── Row 1: group headers ──────────────────────────────────────────────────
    n_fixed  = len(fixed_headers)
    n_vs     = len(vuln_statuses)
    n_hs     = len(HOTSPOT_STATUSES)

    # column ranges (1-based)
    vuln_start   = n_fixed + 1
    vuln_end     = vuln_start + n_vs          # total col + status cols
    hs_start     = vuln_end + 1
    hs_end       = hs_start + n_hs            # total col + status cols

    # Group label cells (merged)
    _hdr_cell(ws, 1, 1, "",              fill=_HDR_FILL)
    if n_fixed > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_fixed)

    _hdr_cell(ws, 1, vuln_start, "VULNERABILITIES", fill=_GRP_FILL)
    if vuln_end > vuln_start:
        ws.merge_cells(start_row=1, start_column=vuln_start,
                       end_row=1, end_column=vuln_end)

    _hdr_cell(ws, 1, hs_start, "SECURITY HOTSPOTS", fill=_GRP_FILL)
    if hs_end > hs_start:
        ws.merge_cells(start_row=1, start_column=hs_start,
                       end_row=1, end_column=hs_end)

    ws.row_dimensions[1].height = 22

    # ── Row 2: column headers ─────────────────────────────────────────────────
    col = 1
    for h in fixed_headers:
        _hdr_cell(ws, 2, col, h); col += 1

    # Vuln columns: total + by status
    _hdr_cell(ws, 2, col, "Total");          col += 1
    for s in vuln_statuses:
        _hdr_cell(ws, 2, col, s.replace("_", " ").title()); col += 1

    # Hotspot columns
    _hdr_cell(ws, 2, col, "Total");          col += 1
    for s in HOTSPOT_STATUSES:
        _hdr_cell(ws, 2, col, s.replace("_", " ").title()); col += 1

    ws.row_dimensions[2].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, item in enumerate(rows, 3):
        fill = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        col  = 1

        for val, al in item["fixed"]:
            _body_cell(ws, row_idx, col, val, fill=fill, align=al); col += 1

        vuln = item["vulns"]
        vfill = _WARN_FILL if vuln.total > 0 else fill
        _body_cell(ws, row_idx, col, vuln.total,    fill=vfill, align=_RIGHT, bold=vuln.total > 0); col += 1
        for s in vuln_statuses:
            v = vuln.by_status.get(s, 0)
            _body_cell(ws, row_idx, col, v or None, fill=vfill, align=_RIGHT); col += 1

        hs  = item["hotspots"]
        hfill = _WARN_FILL if hs.total > 0 else fill
        _body_cell(ws, row_idx, col, hs.total,      fill=hfill, align=_RIGHT, bold=hs.total > 0); col += 1
        for s in HOTSPOT_STATUSES:
            v = hs.by_status.get(s, 0)
            _body_cell(ws, row_idx, col, v or None, fill=hfill, align=_RIGHT); col += 1

        # format date cell
        date_col = item.get("date_col")
        if date_col:
            ws.cell(row=row_idx, column=date_col).number_format = "YYYY-MM-DD HH:MM"

    _auto_width(ws)


def _build_branches_sheet(ws, reports: list[ProjectReport],
                           vuln_statuses: list[str]) -> None:
    fixed_headers = ["Project Name", "Project Key", "Branch", "Type", "Main?", "Last Scan"]
    rows = []
    for r in sorted(reports, key=lambda x: x.name.lower()):
        for b in sorted(r.branches, key=lambda x: (not x.is_main, x.name.lower())):
            rows.append({
                "fixed": [
                    (r.name,                _LEFT),
                    (r.key,                 _LEFT),
                    (b.name,                _LEFT),
                    (b.branch_type,         _CENTER),
                    ("Yes" if b.is_main else "",  _CENTER),
                    (_parse_iso(b.last_analysis), _CENTER),
                ],
                "vulns":    b.vulnerabilities,
                "hotspots": b.hotspots,
                "date_col": 6,
            })
    _write_security_sheet(ws, "Branches", fixed_headers, rows, vuln_statuses)


def _build_pr_sheet(ws, reports: list[ProjectReport],
                    vuln_statuses: list[str]) -> None:
    fixed_headers = ["Project Name", "Project Key", "PR Key", "PR Title",
                     "Source Branch", "Target Branch", "Last Scan"]
    rows = []
    for r in sorted(reports, key=lambda x: x.name.lower()):
        for pr in sorted(r.pull_requests, key=lambda x: x.key):
            rows.append({
                "fixed": [
                    (r.name,                     _LEFT),
                    (r.key,                      _LEFT),
                    (pr.key,                     _CENTER),
                    (pr.title,                   _LEFT),
                    (pr.source_branch,           _LEFT),
                    (pr.target_branch,           _LEFT),
                    (_parse_iso(pr.last_analysis), _CENTER),
                ],
                "vulns":    pr.vulnerabilities,
                "hotspots": pr.hotspots,
                "date_col": 7,
            })
    _write_security_sheet(ws, "Pull Requests", fixed_headers, rows, vuln_statuses)


def export_to_excel(reports: list[ProjectReport], path: str) -> None:
    vuln_statuses, _ = _collect_dynamic_cols(reports)

    wb = Workbook()
    # openpyxl creates a default sheet — rename it to Summary
    _build_summary_sheet(wb.active, reports)
    _build_branches_sheet(wb.create_sheet(), reports, vuln_statuses)

    has_prs = any(r.pull_requests for r in reports)
    if has_prs:
        _build_pr_sheet(wb.create_sheet(), reports, vuln_statuses)

    wb.save(path)


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="SonarQube Security Report — vulnerabilities & hotspots "
                    "across all projects, branches, and PRs.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--url",      default=os.getenv("SONAR_URL"),
                   help="SonarQube base URL  e.g. http://sonar.example.com  [env: SONAR_URL]")
    p.add_argument("--username", default=os.getenv("SONAR_USERNAME"),
                   help="SonarQube username  [env: SONAR_USERNAME]")
    p.add_argument("--password", default=os.getenv("SONAR_PASSWORD"),
                   help="SonarQube password  [env: SONAR_PASSWORD]")
    p.add_argument("--output",   default=os.getenv("SONAR_OUTPUT", "sonar_report.json"),
                   help="Path for the JSON output file  [env: SONAR_OUTPUT]")
    p.add_argument("--excel",    default=os.getenv("SONAR_EXCEL", "sonar_report.xlsx"),
                   help="Path for the Excel (.xlsx) output file  [env: SONAR_EXCEL]")
    p.add_argument("--concurrency", type=int,
                   default=int(os.getenv("SONAR_CONCURRENCY", DEFAULT_CONCURRENCY)),
                   help="Max simultaneous API requests  [env: SONAR_CONCURRENCY]")
    p.add_argument("--summary-only", action="store_true",
                   help="Print only the summary table, skip per-branch detail")
    p.add_argument("--no-verify-ssl", action="store_true",
                   help="Disable SSL certificate verification "
                        "(useful for self-signed certs)")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    missing = [f"--{k}" for k, v in [("url", args.url), ("username", args.username), ("password", args.password)] if not v]
    if missing:
        print(f"Error: missing required values: {', '.join(missing)}\n"
              f"Pass them as CLI args or set SONAR_URL / SONAR_USERNAME / SONAR_PASSWORD in your environment / .env file.",
              file=sys.stderr)
        sys.exit(1)

    print("SonarQube Security Report")
    print(f"  URL         : {args.url}")
    print(f"  Username    : {args.username}")
    print(f"  Concurrency : {args.concurrency}")
    print(f"  JSON output : {args.output}")
    print(f"  Excel output: {args.excel}")
    print(f"  Verify SSL  : {not args.no_verify_ssl}")
    print()

    reports = asyncio.run(run(args))

    print_summary(reports)

    if not args.summary_only:
        print_detail(reports)

    # JSON output
    output = to_dict(reports)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nJSON report saved  → {args.output}")

    # Excel output
    export_to_excel(reports, args.excel)
    print(f"Excel report saved → {args.excel}")

    print(f"\nGRAND TOTALS  ({len(reports)} project(s))")
    print(f"  Vulnerabilities  : {output['grand_total_vulnerabilities']}")
    print(f"  Security Hotspots: {output['grand_total_hotspots']}")


if __name__ == "__main__":
    main()
