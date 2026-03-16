#!/usr/bin/env python3
"""
SonarQube Vulnerability Reopener
=================================
Reopens all non-OPEN, non-CLOSED vulnerabilities across a list of projects
and all their branches.

Statuses transitioned back to OPEN:
  CONFIRMED       → unconfirm
  FALSE_POSITIVE  → reopen
  ACCEPTED        → reopen
  FIXED           → reopen

CLOSED issues are skipped (auto-managed by the scanner).
OPEN issues are not queried at all.

Input:
  A text file with one project key per line.
  Blank lines and lines starting with # are ignored.

Output:
  - Console progress and summary
  - Excel report: Summary sheet (per project+branch) + Detail sheet (per issue)

Usage:
    python sonar_reopen_vulnerabilities.py \
        --url http://sonarqube.example.com \
        --token <user_token> \
        --projects projects.txt

    # Dry run — scans and reports without making any changes
    python sonar_reopen_vulnerabilities.py ... --dry-run

Requirements:
    pip install httpx openpyxl tqdm python-dotenv

Notes:
  - Requires a user token (not a project/analysis token) with
    Administer Issues permission on each target project.
  - SonarQube caps issue pagination at 10,000 results per query.
    Projects/branches exceeding this will log a warning.
"""

import asyncio
import httpx
import argparse
import sys
import os
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Optional, Any
from datetime import datetime, timezone

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from tqdm import tqdm
from tqdm.asyncio import tqdm as atqdm

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ──────────────────────────────────────────────────────────────────

PAGE_SIZE           = 500
DEFAULT_CONCURRENCY = 10

# Statuses to transition back to OPEN → transition name
# Uses the new issueStatus values introduced in SonarQube 10.4 / 2025.1 LTA
TRANSITION_MAP: dict[str, str] = {
    "CONFIRMED":      "unconfirm",
    "FALSE_POSITIVE": "reopen",
    "ACCEPTED":       "reopen",
    "FIXED":          "reopen",
}

STATUSES_TO_FETCH = ",".join(TRANSITION_MAP.keys())


# ── Data Models ────────────────────────────────────────────────────────────────

@dataclass
class IssueToProcess:
    issue_key:      str
    project_key:    str
    project_name:   str
    branch:         str
    message:        str
    component:      str
    severity:       str
    current_status: str
    transition:     str


@dataclass
class IssueResult:
    issue:   IssueToProcess
    outcome: str            # "transitioned" | "failed" | "dry_run"
    error:   Optional[str] = None


# ── SonarQube Client ───────────────────────────────────────────────────────────

class SonarQubeClient:
    def __init__(self, base_url: str, token: str,
                 concurrency: int = DEFAULT_CONCURRENCY,
                 verify_ssl: bool = True):
        self.base_url = base_url.rstrip("/")
        self._headers = {"Authorization": f"Bearer {token}"}
        self._sem     = asyncio.Semaphore(concurrency)

    async def _get(self, client: httpx.AsyncClient, path: str,
                   params: dict[str, Any] | None = None) -> dict:
        async with self._sem:
            resp = await client.get(
                f"{self.base_url}{path}",
                params=params or {},
                headers=self._headers,
                timeout=60.0,
            )
            resp.raise_for_status()
            return resp.json()

    async def _post(self, client: httpx.AsyncClient, path: str,
                    data: dict[str, Any] | None = None) -> dict:
        async with self._sem:
            resp = await client.post(
                f"{self.base_url}{path}",
                data=data or {},
                headers=self._headers,
                timeout=60.0,
            )
            resp.raise_for_status()
            return resp.json()

    async def get_project_name(self, client: httpx.AsyncClient,
                                project_key: str) -> str:
        try:
            data = await self._get(client, "/api/projects/search",
                                   {"projects": project_key, "ps": 1})
            comps = data.get("components", [])
            return comps[0]["name"] if comps else project_key
        except Exception:
            return project_key

    async def get_branches(self, client: httpx.AsyncClient,
                           project_key: str) -> list[dict]:
        data = await self._get(client, "/api/project_branches/list",
                               {"project": project_key})
        return data.get("branches", [])

    async def get_issues_to_reopen(self, client: httpx.AsyncClient,
                                    project_key: str,
                                    branch: str) -> list[dict]:
        """Paginate all non-OPEN, non-CLOSED vulnerabilities for a branch."""
        issues: list[dict] = []
        page = 1
        while True:
            data = await self._get(client, "/api/issues/search", {
                "componentKeys":           project_key,
                "impactSoftwareQualities": "SECURITY",
                "issueStatuses":           STATUSES_TO_FETCH,
                "branch":                  branch,
                "ps":                      PAGE_SIZE,
                "p":                       page,
            })
            batch = data.get("issues", [])
            issues.extend(batch)
            pg    = data.get("paging", {})
            total = pg.get("total", 0)
            if page == 1 and total > 10_000:
                _warn(f"[{project_key}/{branch}] {total} issues found; "
                      f"SonarQube caps pagination at 10,000 — only first 10,000 processed.")
            if pg.get("pageIndex", 1) * pg.get("pageSize", PAGE_SIZE) >= min(total, 10_000):
                break
            page += 1
        return issues

    async def do_transition(self, client: httpx.AsyncClient,
                             issue_key: str, transition: str) -> None:
        await self._post(client, "/api/issues/do_transition", {
            "issue":      issue_key,
            "transition": transition,
        })


# ── Helpers ────────────────────────────────────────────────────────────────────

def _warn(msg: str) -> None:
    tqdm.write(f"  [WARN] {msg}", file=sys.stderr)


def _severity(issue: dict) -> str:
    """Return severity string — prefers legacy field, falls back to impacts."""
    if sev := issue.get("severity"):
        return sev
    impacts = issue.get("impacts", [])
    return impacts[0].get("severity", "") if impacts else ""


def load_project_keys(path: str) -> list[str]:
    with open(path, encoding="utf-8") as f:
        return [
            line.strip()
            for line in f
            if line.strip() and not line.strip().startswith("#")
        ]


# ── Orchestration ──────────────────────────────────────────────────────────────

async def scan_branch(sonar: SonarQubeClient, client: httpx.AsyncClient,
                       project_key: str, project_name: str,
                       branch: str, pbar: tqdm) -> list[IssueToProcess]:
    try:
        raw = await sonar.get_issues_to_reopen(client, project_key, branch)
        result: list[IssueToProcess] = []
        for issue in raw:
            # Prefer new issueStatus field (2025.1 LTA), fall back to status
            status = issue.get("issueStatus") or issue.get("status", "")
            if status not in TRANSITION_MAP:
                continue
            result.append(IssueToProcess(
                issue_key=issue["key"],
                project_key=project_key,
                project_name=project_name,
                branch=branch,
                message=issue.get("message", ""),
                component=issue.get("component", ""),
                severity=_severity(issue),
                current_status=status,
                transition=TRANSITION_MAP[status],
            ))
        return result
    except Exception as exc:
        _warn(f"Failed to scan {project_key}/{branch}: {exc}")
        return []
    finally:
        pbar.update(1)


async def reopen_issue(sonar: SonarQubeClient, client: httpx.AsyncClient,
                        issue: IssueToProcess, dry_run: bool,
                        pbar: tqdm) -> IssueResult:
    if dry_run:
        pbar.update(1)
        return IssueResult(issue=issue, outcome="dry_run")
    try:
        await sonar.do_transition(client, issue.issue_key, issue.transition)
        pbar.update(1)
        return IssueResult(issue=issue, outcome="transitioned")
    except Exception as exc:
        pbar.update(1)
        return IssueResult(issue=issue, outcome="failed", error=str(exc))


async def run(args: argparse.Namespace) -> list[IssueResult]:
    project_keys = load_project_keys(args.projects)
    if not project_keys:
        print("No project keys found in input file.", file=sys.stderr)
        sys.exit(1)

    sonar = SonarQubeClient(
        base_url=args.url,
        token=args.token,
        concurrency=args.concurrency,
        verify_ssl=not args.no_verify_ssl,
    )

    async with httpx.AsyncClient(verify=not args.no_verify_ssl) as client:

        # ── Phase 1: resolve project names + branches ──────────────────────────
        async def fetch_project_info(key: str) -> dict:
            name, branches = await asyncio.gather(
                sonar.get_project_name(client, key),
                sonar.get_branches(client, key),
            )
            return {"key": key, "name": name, "branches": branches}

        projects: list[dict] = await atqdm.gather(
            *[fetch_project_info(k) for k in project_keys],
            desc="Fetching project info",
            unit="project",
            total=len(project_keys),
        )

        total_branches = sum(len(p["branches"]) for p in projects)
        tqdm.write(f"\n{len(projects)} project(s), {total_branches} branch(es) total.\n")

        # ── Phase 2: scan all branches for issues to reopen ────────────────────
        with tqdm(desc="Scanning branches", unit="branch",
                  total=total_branches) as pbar:
            scan_results = await asyncio.gather(*[
                scan_branch(sonar, client, p["key"], p["name"], b["name"], pbar)
                for p in projects
                for b in p["branches"]
            ])

        all_issues: list[IssueToProcess] = [
            issue for batch in scan_results for issue in batch
        ]
        tqdm.write(f"\nFound {len(all_issues)} issue(s) to reopen.\n")
        if not all_issues:
            return []

        # ── Phase 3: reopen issues ─────────────────────────────────────────────
        label = "Dry run" if args.dry_run else "Reopening"
        with tqdm(desc=f"{label} issues", unit="issue",
                  total=len(all_issues)) as pbar:
            results: list[IssueResult] = list(await asyncio.gather(*[
                reopen_issue(sonar, client, issue, args.dry_run, pbar)
                for issue in all_issues
            ]))

    return results


# ── Console summary ────────────────────────────────────────────────────────────

def print_summary(results: list[IssueResult]) -> None:
    counts = Counter(r.outcome for r in results)
    print("\n" + "═" * 60)
    print("  REOPEN SUMMARY")
    print("═" * 60)
    print(f"  Total processed  : {len(results)}")
    print(f"  Transitioned     : {counts.get('transitioned', 0)}")
    print(f"  Failed           : {counts.get('failed', 0)}")
    print(f"  Dry run          : {counts.get('dry_run', 0)}")
    print("═" * 60)


# ── Excel export ───────────────────────────────────────────────────────────────

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_OK_FILL    = PatternFill("solid", fgColor="E2EFDA")   # green  — transitioned
_FAIL_FILL  = PatternFill("solid", fgColor="FCE4D6")   # orange — failed
_DRY_FILL   = PatternFill("solid", fgColor="DEEAF1")   # blue   — dry run
_ALT_FILL   = PatternFill("solid", fgColor="F2F2F2")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="375623")

_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT  = Font(name="Calibri", size=10)
_BOLD_FONT  = Font(bold=True, name="Calibri", size=10)
_TOTAL_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

_THIN   = Side(border_style="thin", color="B8CCE4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")


def _hdr(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _HDR_FILL; c.font = _HDR_FONT
    c.alignment = _CENTER; c.border = _BORDER


def _cell(ws, row: int, col: int, value,
          fill=None, align=None, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill or _WHITE_FILL
    c.font = _BOLD_FONT if bold else _BODY_FONT
    c.alignment = align or _LEFT; c.border = _BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _outcome_fill(outcome: str) -> PatternFill:
    return {"transitioned": _OK_FILL, "failed": _FAIL_FILL,
            "dry_run": _DRY_FILL}.get(outcome, _WHITE_FILL)


def _build_summary_sheet(ws, results: list[IssueResult]) -> None:
    ws.title = "Summary"
    ws.freeze_panes = "A2"

    headers = ["Project Key", "Project Name", "Branch",
               "Transitioned", "Failed", "Dry Run", "Total"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    stats: dict[tuple, dict] = defaultdict(
        lambda: {"name": "", "t": 0, "f": 0, "d": 0}
    )
    for r in results:
        k = (r.issue.project_key, r.issue.branch)
        stats[k]["name"] = r.issue.project_name
        if r.outcome == "transitioned": stats[k]["t"] += 1
        elif r.outcome == "failed":     stats[k]["f"] += 1
        elif r.outcome == "dry_run":    stats[k]["d"] += 1

    for row_idx, ((proj_key, branch), s) in enumerate(sorted(stats.items()), 2):
        fill  = _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL
        total = s["t"] + s["f"] + s["d"]
        _cell(ws, row_idx, 1, proj_key,  fill=fill)
        _cell(ws, row_idx, 2, s["name"], fill=fill)
        _cell(ws, row_idx, 3, branch,    fill=fill)
        _cell(ws, row_idx, 4, s["t"],    fill=fill, align=_RIGHT)
        _cell(ws, row_idx, 5, s["f"],    fill=fill, align=_RIGHT)
        _cell(ws, row_idx, 6, s["d"],    fill=fill, align=_RIGHT)
        _cell(ws, row_idx, 7, total,     fill=fill, align=_RIGHT, bold=True)

    total_row = len(stats) + 2
    t = sum(s["t"] for s in stats.values())
    f = sum(s["f"] for s in stats.values())
    d = sum(s["d"] for s in stats.values())
    _cell(ws, total_row, 1, f"TOTAL  ({len(stats)} branch(es))",
          fill=_TOTAL_FILL, bold=True)
    ws.cell(row=total_row, column=1).font = _TOTAL_FONT
    for col in [2, 3]:
        ws.cell(row=total_row, column=col).fill   = _TOTAL_FILL
        ws.cell(row=total_row, column=col).border = _BORDER
    for col, val in [(4, t), (5, f), (6, d), (7, t + f + d)]:
        _cell(ws, total_row, col, val, fill=_TOTAL_FILL, align=_RIGHT, bold=True)
        ws.cell(row=total_row, column=col).font = _TOTAL_FONT

    _auto_width(ws)


def _build_detail_sheet(ws, results: list[IssueResult]) -> None:
    ws.title = "Detail"
    ws.freeze_panes = "A2"

    headers = ["Project Key", "Project Name", "Branch", "Issue Key",
               "Component", "Message", "Severity",
               "Previous Status", "Transition Applied", "Outcome", "Error"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 30

    for row_idx, r in enumerate(results, 2):
        fill = _outcome_fill(r.outcome)
        i    = r.issue
        _cell(ws, row_idx,  1, i.project_key,    fill=fill)
        _cell(ws, row_idx,  2, i.project_name,   fill=fill)
        _cell(ws, row_idx,  3, i.branch,         fill=fill)
        _cell(ws, row_idx,  4, i.issue_key,      fill=fill)
        _cell(ws, row_idx,  5, i.component,      fill=fill)
        _cell(ws, row_idx,  6, i.message,        fill=fill)
        _cell(ws, row_idx,  7, i.severity,       fill=fill, align=_CENTER)
        _cell(ws, row_idx,  8, i.current_status, fill=fill, align=_CENTER)
        _cell(ws, row_idx,  9, i.transition,     fill=fill, align=_CENTER)
        _cell(ws, row_idx, 10, r.outcome,        fill=fill, align=_CENTER, bold=True)
        _cell(ws, row_idx, 11, r.error or "",    fill=fill)

    _auto_width(ws)


def export_excel(results: list[IssueResult], path: str) -> None:
    wb = Workbook()
    _build_summary_sheet(wb.active, results)
    _build_detail_sheet(wb.create_sheet(), results)
    wb.save(path)


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Reopen SonarQube vulnerabilities across specified projects and all branches.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--url",      default=os.getenv("SONAR_URL"),
                   help="SonarQube base URL  [env: SONAR_URL]")
    p.add_argument("--token",    default=os.getenv("SONAR_TOKEN"),
                   help="SonarQube user token with Administer Issues permission  [env: SONAR_TOKEN]")
    p.add_argument("--projects", default=os.getenv("SONAR_PROJECTS_FILE", "projects.txt"),
                   help="Text file with project keys, one per line  [env: SONAR_PROJECTS_FILE]")
    p.add_argument("--output",   default=os.getenv("SONAR_REOPEN_OUTPUT", "reopen_report.xlsx"),
                   help="Excel report output path  [env: SONAR_REOPEN_OUTPUT]")
    p.add_argument("--concurrency", type=int,
                   default=int(os.getenv("SONAR_CONCURRENCY", DEFAULT_CONCURRENCY)),
                   help="Max parallel API requests  [env: SONAR_CONCURRENCY]")
    p.add_argument("--dry-run", action="store_true",
                   help="Scan without making any changes — report what would be transitioned")
    p.add_argument("--no-verify-ssl", action="store_true",
                   help="Disable SSL certificate verification (self-signed certs)")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    missing = [f"--{k}" for k, v in
               [("url", args.url), ("token", args.token)] if not v]
    if missing:
        print(f"Error: missing required values: {', '.join(missing)}\n"
              f"Pass them as CLI args or set SONAR_URL / SONAR_TOKEN "
              f"in your environment / .env file.",
              file=sys.stderr)
        sys.exit(1)

    print("SonarQube Vulnerability Reopener")
    print(f"  URL         : {args.url}")
    print(f"  Projects    : {args.projects}")
    print(f"  Concurrency : {args.concurrency}")
    print(f"  Output      : {args.output}")
    print(f"  Dry run     : {args.dry_run}")
    print(f"  Verify SSL  : {not args.no_verify_ssl}")
    print()

    results = asyncio.run(run(args))

    if not results:
        print("No issues found to reopen.")
        return

    print_summary(results)
    export_excel(results, args.output)
    print(f"\nExcel report saved → {args.output}")


if __name__ == "__main__":
    main()
